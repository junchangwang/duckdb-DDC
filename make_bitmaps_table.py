#!/usr/bin/env python3
"""Bitmaps for TPC-H — SF10 storage table.

数字全部来自 [load_bitmap_breakdown] 真实测量。
配色：黑白，关键句子用红色。
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUT = "/home/lichenhang/lee/thesis/check/lee/duckdb-dev/Bitmaps_for_TPCH_SF10.xlsx"

# ---------------------------------------------------------------------------
# Comments
# ---------------------------------------------------------------------------
# 写法：每个 Q 一段
#   Q<N>
#   --
#   pipeline 步骤，-> 表示下一步，+ 表示同一步多个输入
# [RED]...[/RED] 标记的句子在 Excel 里渲染成红色加粗
# ---------------------------------------------------------------------------

C_RETURNFLAG = """\
Q1
--
shipdate ≤ 1998-09-02 反取（约 90 天 OR 后 NOT）
-> AND linestatus(k=2) AND returnflag(k=3)
-> 拼出 (rf, ls) 笛卡尔积 6 个 group mask，转 MSB byte stream
-> 顺扫 60M 行 lineitem，每行 5 个 bit-test 判断属于哪一组
-> per-row SUM(qty / extprice / disc / charge) + COUNT

Q10
--
orders [1993-10, 1994-01) 扫
-> ~1500 个 orderkey 做 k-way OR + 单键 OR returnflag='R'
-> AND -> get_rowids -> BMFetch
+ per-custkey revenue 累加
-> top-20 heap"""

C_LINESTATUS = """\
Q1
--
shipdate ≤ 1998-09-02 反取（约 90 天 OR 后 NOT）
-> AND linestatus(k=2) AND returnflag(k=3)
-> 拼出 (rf, ls) 笛卡尔积 6 个 group mask，转 MSB byte stream
-> 顺扫 60M 行 lineitem，每行 5 个 bit-test 判断属于哪一组
-> per-row SUM(qty / extprice / disc / charge) + COUNT"""

C_SHIPDATE = """\
Q1
--
shipdate > 1998-09-02 做约 90 天 range OR -> NOT
-> 和 linestatus + returnflag 一起拼 group mask
-> 顺扫 60M 行 + per-row 聚合

Q3
--
shipdate > 1995-03-15 做 200 天 range OR
-> AND orderkey OR(~1500 key)
-> get_rowids -> BMFetch
+ per-orderkey revenue
-> top-10 heap

Q14
--
shipdate ∈ [1995-09-01, 1995-10-01) 做 30 天 range OR
-> get_rowids
+ part 表扫 p_type LIKE 'PROMO%' -> partkey 集合
-> BMFetch lineitem
+ per-row 判断 partkey 是否在集合里
-> promo / total revenue 比

Q15
--
shipdate ∈ [1996-01-01, 1996-04-01) 做 91 天 range OR
-> 顺扫 BMFetch (suppkey, price, disc)
+ per-row 按 suppkey 累加 revenue
-> 找 max revenue + 并列 suppkey 集合"""

C_SHIPDATE_BPE = """\
Q1 / Q3 BPE 加速路径（β 方案）
--
range 查询拆成 prefix(LAST) AND_NOT prefix(B)
+ per-day 边界 OR over (cutoff, bucket_max(B)]
-> O(1) bitmap 操作代替原来 O(k) 个 per-day OR
[RED]只有 DEBIT_BM=cb_bpe / cr_bpe 时才会构建这张表[/RED]"""

C_SHIPDATE_GE = """\
Q6 Group Encoding，整张表只做 1 次 OR
--
shipdate_GE.apply_or_to(dst, 1994) -> 单键 mask
-> AND discount[5,7] (3 OR)
-> AND quantity<24 (24 OR)
-> get_rowids -> BMFetch (disc, price) -> SUM
[RED]用 1 张年级别 bitmap 代替 365 天 per-day OR[/RED]"""

C_ORDERKEY = """\
Q3 (~1500 key)
--
customer 扫 'BUILDING' + orders [<1995-03-15] 扫
-> matched orderkey 做 k-way OR（DDC 用 sca scatter）
-> AND shipdate 200 天 range -> BMFetch
+ per-orderkey revenue -> top-10 heap

Q4 (~800 key)
--
orders [1993-07, 1993-10) 扫 -> matched orderkey
-> orderkey k-way OR -> get_rowids -> BMFetch lineitem
+ per-row commit<receipt EXISTS 判断
-> per-priority COUNT

Q5 (~590K key，全 TPC-H 最大的 k-way OR)
--
region -> nation -> customer -> orders -> supplier 五表半连接
-> orderkey OR + suppkey OR
-> AND -> get_rowids -> BMFetch
+ per-row (okey 国家 == skey 国家) 校验
+ per-nation revenue 聚合

Q8 (~万级 key)
--
region -> nation -> customer + supplier 半连接 + part 扫
-> orderkey OR + partkey OR
-> AND -> BMFetch
+ per-row 年份提取 + (国家 == BRAZIL ?) 分子分母
-> per-year mkt_share

Q10 (~1500 key)
--
orders [1993-10, 1994-01) 扫
-> orderkey OR + 单键 OR returnflag='R'
-> AND -> BMFetch
+ per-custkey revenue -> top-20 heap"""

C_SUPPKEY = """\
Q5 (~8K key)
--
region -> nation -> customer / supplier 半连接 -> ~8K suppkey
-> DDC::or_many（sca scatter_or_decompressed）
-> AND orderkey OR (~590K) -> get_rowids -> BMFetch
+ per-row (okey 国家 == skey 国家) 校验
+ per-nation revenue 聚合"""

C_PARTKEY = """\
Q8 (~2K key)
--
part 扫 p_type='ECONOMY ANODIZED STEEL'
-> matched partkey 做 k-way OR
-> AND orderkey OR -> BMFetch
+ per-row 年份提取 + per-nation volume
-> BRAZIL / total mkt_share

Q17 (~200 key)
--
part 扫 p_brand='#23' AND p_container='MED BOX'
-> ~200 matched partkey 做 batched OR -> BMFetch
+ per-partkey 算 avg(qty)，BMFetch 结果缓存复用
+ per-row 判断 qty < 0.2 * avg(qty)
-> SUM(price) / 7.0"""

C_DISCOUNT = """\
Q6 range OR [5,7]，3 个 OR
--
discount.apply_or_range_to(dst, 5, 7) -> l_discount ∈ [0.05, 0.07]
-> AND shipdate_GE(1994) AND quantity<24
-> get_rowids -> BMFetch (disc, price) -> SUM"""

C_DISCOUNT_BPE = """\
Q6 BPE 加速（β 方案，bucket=1 与 per-value 完全对齐）
--
disc ∈ [5,7] = prefix(7) AND_NOT prefix(4)
-> O(1) bitmap 操作代替 3 个 OR
[RED]只有 DEBIT_BM=cb_bpe / cr_bpe 时才会构建这张表[/RED]"""

C_QUANTITY = """\
Q6 range OR [0, 2400)，24 个 OR
--
quantity.apply_or_range_to(dst, 0, 2399) -> l_quantity < 24.00
-> AND shipdate_GE(1994) AND discount[5,7]
-> get_rowids -> BMFetch (disc, price) -> SUM"""

C_QUANTITY_BPE = """\
Q6 BPE 加速（β 方案，bucket=500）
--
qty ≤ 2399 = prefix(bucket_of(2399))
+ 边界 raw ∈ (2399, bucket_max] 走 per-value 修补
-> O(1) bitmap + 边界修正 代替 24 个 OR
[RED]只有 DEBIT_BM=cb_bpe / cr_bpe 时才会构建这张表[/RED]"""

C_SHIPMODE = """\
Q12 (双键 {'M','S'})
--
shipmode k=2 OR（VARCHAR 取首字节 ASCII 作 key）
-> AND receiptdate 365 天 range OR
-> get_rowids -> BMFetch lineitem
+ per-row ship<commit<receipt 双重比较
+ per-orderkey priority 查表 -> (mode, class) 计数

Q19 (单键 'A')
--
shipmode='A' (k=1，'A' 同时覆盖 AIR + AIR REG)
-> AND shipinstruct='D' (k=1)
-> get_rowids
+ part 扫 brand+container+size 建 group spec 表
-> BMFetch lineitem
+ 3 个 group 分类 + per-row qty range 判断
-> revenue 聚合"""

C_SHIPINSTRUCT = """\
Q19 (单键 'D')
--
shipinstruct='D' (k=1，DELIVER IN PERSON)
-> AND shipmode='A' (k=1)
-> get_rowids -> part 扫 -> BMFetch
+ 3 个 group 分类 + per-row qty range 判断
-> revenue 聚合"""

C_RECEIPTDATE = """\
Q12 (365 天 range OR)
--
receiptdate ∈ [1994-01-01, 1995-01-01) -> 365 天 OR
-> AND shipmode 'M' / 'S' (k=2)
-> get_rowids -> BMFetch
+ per-row ship<commit<receipt 双重比较
+ per-orderkey priority 查表 -> (mode, class) 计数"""

# ---------------------------------------------------------------------------
ROWS = [
    ("returnflag",      3,            C_RETURNFLAG,
     23.22, 20.20, 20.20, 22.53, 22.53, 22.53, 22.49, None, None, 23.21),
    ("linestatus",      2,            C_LINESTATUS,
     15.48, 14.25, 14.25, 15.02, 15.02, 15.02, 15.00, None, None, 15.48),
    ("shipdate",        2526,         C_SHIPDATE,
     466.97, 258.05, 258.05, 138.44, 138.44, 138.44, 921.39, None, None, 239.83),
    ("  shipdate_BPE",  "26 buckets", C_SHIPDATE_BPE,
     None, None, 194.96, None, 189.24, None, None, None, None, None),
    ("shipdate_GE",     7,            C_SHIPDATE_GE,
     50.85, 26.48, 26.48, 52.56, 52.56, 52.52, 52.26, None, None, 49.82),
    ("orderkey",        15_000_000,   C_ORDERKEY,
     305.80, 96.45, 96.45, 359.98, 359.98, 257.15, 365.62, None, None, 185.80),
    ("suppkey",         100_000,      C_SUPPKEY,
     480.94, 356.11, 356.11, 472.84, 472.84, 472.84, 959.97, None, None, 241.14),
    ("partkey",         2_000_000,    C_PARTKEY,
     503.88, 359.72, 359.72, 608.09, 608.09, 608.09, 975.75, None, None, 263.94),
    ("discount",        11,           C_DISCOUNT,
     84.91, 54.79, 54.79, 82.57, 82.57, 82.57, 82.48, None, None, 84.19),
    ("  discount_BPE",  "11 buckets", C_DISCOUNT_BPE,
     None, None, 82.48, None, 75.13, None, None, None, None, None),
    ("quantity",        50,           C_QUANTITY,
     276.43, 97.56, 97.56, 120.34, 120.34, 120.34, 346.69, None, None, 206.44),
    ("  quantity_BPE",  "11 buckets", C_QUANTITY_BPE,
     None, None, 82.48, None, 82.62, None, None, None, None, None),
    ("shipmode",        7,            C_SHIPMODE,
     46.44, 38.97, 38.97, 45.06, 45.06, 45.06, 44.99, None, None, 46.42),
    ("shipinstruct",    4,            C_SHIPINSTRUCT,
     30.96, 29.77, 29.77, 30.04, 30.04, 30.04, 29.99, None, None, 30.96),
    ("receiptdate",     2526,         C_RECEIPTDATE,
     467.57, 258.33, 258.33, 138.52, 138.52, 138.52, 922.61, None, None, 239.84),
]

ALGO_COLS = ["WAH", "DDC", "DDC+BPE", "CRoaring", "CRoaring+BPE",
             "CRoaring+Run", "EWAH", "Bitset", "Bitset+AVX512", "Concise"]

# ---------------------------------------------------------------------------
# Rich-text comment builder: parse [RED]...[/RED] markers.
# ---------------------------------------------------------------------------
BASE_INLINE = InlineFont(rFont="Calibri", sz=9, color="000000")
RED_INLINE  = InlineFont(rFont="Calibri", sz=9, color="C00000", b=True)

def build_rich(text: str) -> CellRichText:
    blocks = []
    pos = 0
    while pos < len(text):
        s = text.find("[RED]", pos)
        if s == -1:
            if pos < len(text):
                blocks.append(TextBlock(BASE_INLINE, text[pos:]))
            break
        if s > pos:
            blocks.append(TextBlock(BASE_INLINE, text[pos:s]))
        e = text.find("[/RED]", s)
        if e == -1:
            blocks.append(TextBlock(BASE_INLINE, text[s:]))
            break
        blocks.append(TextBlock(RED_INLINE, text[s + 5:e]))
        pos = e + 6
    return CellRichText(blocks)

# ---------------------------------------------------------------------------
# Workbook
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Bitmaps for TPC-H"

# Black & white palette
HEADER_FILL = PatternFill("solid", fgColor="000000")  # black
TOTAL_FILL  = PatternFill("solid", fgColor="000000")  # black
BPE_FILL    = PatternFill("solid", fgColor="D9D9D9")  # medium gray
BAND_FILL   = PatternFill("solid", fgColor="F2F2F2")  # very light gray (zebra)

TITLE_FONT  = Font(name="Calibri", size=14, color="FFFFFF", bold=True)
HDR_FONT    = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
BODY        = Font(name="Calibri", size=10, color="000000")
ATTR_FONT   = Font(name="Calibri", size=11, color="000000", bold=True)
BPE_FONT    = Font(name="Calibri", size=10, color="000000", italic=True)
COMMENT_FONT = Font(name="Calibri", size=9, color="000000")
TOTAL_FONT  = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
NA_FONT     = Font(name="Calibri", size=11, color="A6A6A6")

thin   = Side(border_style="thin",   color="000000")
medium = Side(border_style="medium", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK_TOP = Border(left=thin, right=thin, top=medium, bottom=medium)

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTR_T = Alignment(horizontal="center", vertical="top",    wrap_text=True)
LEFT_T = Alignment(horizontal="left",  vertical="top",    wrap_text=True)
RIGHT_T = Alignment(horizontal="right", vertical="top",   wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")

# ----- Title row -----
n_cols = 3 + len(ALGO_COLS)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
tc = ws.cell(row=1, column=1, value="Bitmaps for TPC-H  (SF10, sizes in MB)")
tc.font = TITLE_FONT
tc.fill = HEADER_FILL
tc.alignment = CTR
ws.row_dimensions[1].height = 30

# ----- Header row -----
HEADERS = ["Attribute", "Cardinality", "Comment"] + ALGO_COLS
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font = HDR_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CTR
    cell.border = BORDER
ws.row_dimensions[2].height = 32

# ----- Data rows -----
START = 3
for i, row in enumerate(ROWS):
    r = START + i
    attr, card, comment, *nums = row
    is_bpe = attr.startswith("  ")
    band = (i % 2 == 1) and not is_bpe
    fill = BPE_FILL if is_bpe else (BAND_FILL if band else None)

    # Attribute
    c = ws.cell(row=r, column=1, value=attr)
    c.alignment = LEFT_T
    c.border = BORDER
    c.font = BPE_FONT if is_bpe else ATTR_FONT
    if fill:
        c.fill = fill

    # Cardinality
    c = ws.cell(row=r, column=2, value=card)
    c.alignment = CTR_T
    c.border = BORDER
    c.font = BPE_FONT if is_bpe else BODY
    if isinstance(card, int):
        c.number_format = "#,##0"
    if fill:
        c.fill = fill

    # Comment (rich text — [RED] tagged spans become red bold)
    c = ws.cell(row=r, column=3)
    c.value = build_rich(comment)
    c.alignment = LEFT_T
    c.border = BORDER
    c.font = COMMENT_FONT
    if fill:
        c.fill = fill

    # Algorithm sizes
    for j, v in enumerate(nums):
        col = 4 + j
        cell = ws.cell(row=r, column=col, value=("—" if v is None else v))
        cell.alignment = RIGHT_T if v is not None else CTR_T
        cell.border = BORDER
        if v is None:
            cell.font = NA_FONT
        else:
            cell.font = BODY
            cell.number_format = "#,##0.00"
        if fill:
            cell.fill = fill

    # Dynamic row height — fit comment lines.
    n_lines = comment.count("\n") + 1
    ws.row_dimensions[r].height = max(28, 13 * n_lines + 6)

# ----- Total row -----
total_row = START + len(ROWS)
c = ws.cell(row=total_row, column=1, value="Total")
c.font = TOTAL_FONT
c.fill = TOTAL_FILL
c.alignment = CTR
c.border = THICK_TOP
for col in (2, 3):
    cc = ws.cell(row=total_row, column=col)
    cc.fill = TOTAL_FILL
    cc.border = THICK_TOP

first, last = START, total_row - 1
for j, _ in enumerate(ALGO_COLS):
    col = 4 + j
    L = get_column_letter(col)
    cell = ws.cell(row=total_row, column=col, value=f"=SUM({L}{first}:{L}{last})")
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    cell.alignment = RIGHT
    cell.number_format = "#,##0.00"
    cell.border = THICK_TOP
ws.row_dimensions[total_row].height = 30

# ----- Column widths -----
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 13
ws.column_dimensions["C"].width = 75
for j, name in enumerate(ALGO_COLS):
    w = 14 if "+" in name or len(name) > 6 else 12
    ws.column_dimensions[get_column_letter(4 + j)].width = w

# Freeze: header + first three cols (Attribute / Cardinality / Comment) visible.
ws.freeze_panes = "D3"

# ----- Footer notes -----
notes_row = total_row + 2
ws.merge_cells(start_row=notes_row, start_column=1,
               end_row=notes_row, end_column=n_cols)
nc = ws.cell(row=notes_row, column=1, value=(
    "数据来源：SF10 真实测量，每个数字 = indexed_bitmap.hpp 里 "
    "compute_layers() 对应后端原生格式的压缩字节数。  "
    "灰色行（_BPE 结尾）= 同一列上额外加的累积前缀辅助结构，"
    "只在 DEBIT_BM=cb_bpe / cr_bpe 模式下构建。  "
    "Bitset / Bitset+AVX512 没有实测——uncompressed 每张 bitmap "
    "约 7.5 MB × k，例如 orderkey k=15M 就 ≈ 112 TB，高基数属性物理上不可行。  "
    "CRoaring+Run 把 orderkey（FK，连续 run）从 359.98 → 257.15 MB，"
    "但 shipdate（per-day，无 run）保持 138.44 MB 不变。"))
nc.font = Font(name="Calibri", size=9, italic=True, color="595959")
nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.row_dimensions[notes_row].height = 90

# ----- Save -----
wb.save(OUT)
print(f"[OK] wrote {OUT}")
