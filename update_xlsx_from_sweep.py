#!/usr/bin/env python3
"""
Update Bitmaps_for_TPCH_SF10.xlsx and bm_results.xlsx from the
bm_logs_fix_verify/ sweep output.

Only the Qs we re-measured are touched (Q1, Q10, Q14, Q15, Q17, Q19).
Cells that map to a backend column we did NOT measure (CB+BPE, CR+BPE,
BS, BSA, DuckDB SQL) are left untouched.  Sheets that contain
descriptive text (README, Correctness validation messages) are also
left untouched except for row counts where applicable.
"""
import csv
import re
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl as xl

ROOT       = Path(__file__).resolve().parent
LOG_DIR    = ROOT / "bm_logs_fix_verify"
BITMAPS_XL = ROOT / "Bitmaps_for_TPCH_SF10.xlsx"
RESULTS_XL = ROOT / "bm_results.xlsx"

QS_TO_UPDATE = {1, 10, 14, 15, 17, 19}

# Qs where the BPE column is algorithmically identical to the non-BPE
# column (no range-OR fast-path inside the Q; the load just builds a
# different bitmap class but Q dispatch falls back to the same code).
# For these Qs we mirror CB→CB+BPE and CR→CR+BPE so the BPE column is
# not left stale.  Q1 is the exception: it actually exercises the BPE
# fast-path on shipdate complement OR, so we need a separate Q1 BPE
# re-run for CB+BPE / CR+BPE values.
BPE_MIRRORS_NONBPE = {10, 14, 15, 17, 19}

# (pass, backend-from-sweep) -> xlsx column header.  The xlsx only has
# CB+GE and CR+GE GE-columns; CRR/WAH/EW/CON GE numbers exist in the
# raw logs but are not in the spreadsheet schema, so we drop them.
def col_for(pass_, bm):
    if pass_ == "day":
        return {"cb": "CB", "cr": "CR", "crr": "CRR",
                "wah": "WAH", "ew": "EW", "con": "CON"}.get(bm)
    if pass_ == "ge":
        if bm == "cb_ge": return "CB+GE"
        if bm == "cr":    return "CR+GE"
    return None

# Backend label appearing inside per-iter / summary lines.  Used to
# distinguish per-backend output sections in Q15's combined log.
# Note: Q15 prints SHORT tags (CB/CR/CRR/WAH/EW/CON) regardless of
# cb vs cb_ge, since it always selects run_cb() for both.  Other Qs
# print the IndexedX class name ("DDC", "DDCGE", "CRoaring"…).
Q15_SHORT_LABEL = {
    "cb": "CB", "cb_ge": "CB",
    "cr": "CR", "crr": "CRR",
    "wah": "WAH", "ew": "EW", "con": "CON",
}
BACKEND_LABEL_FOR = {
    "cb": "DDC", "cb_ge": "DDCGE",
    "cr": "CRoaring", "crr": "CRoaringRun",
    "wah": "WAH", "ew": "EWAH", "con": "Concise",
}

# Bitmap loads per Q — must match run_fix_verify_sweep.sh.
LOADS_DAY = {
    1:  ["shipdate", "linestatus", "returnflag"],
    10: ["orderkey", "returnflag"],
    14: ["shipdate"],
    15: ["shipdate"],
    17: ["partkey"],
    19: ["shipmode", "shipinstruct"],
}
LOADS_GE  = {
    1:  ["shipdate_GE_month", "linestatus", "returnflag"],
    10: ["orderkey", "returnflag"],
    14: ["shipdate_GE_month"],
    15: ["suppkey", "shipdate_GE_month"],
    17: ["partkey"],
    19: ["shipmode", "shipinstruct"],
}


def parse_summary_csv(path):
    """
    {(Q, xlsx_col): median_ms} from SUMMARY.csv.  Rows with '-' as the
    median (shutdown-crash cases) are dropped here — the caller fills
    them in from per-iter Total= lines parsed directly from the log.
    """
    out = {}
    with path.open() as f:
        for row in csv.DictReader(f):
            try: q = int(row["q"])
            except: continue
            if q not in QS_TO_UPDATE: continue
            col = col_for(row["pass"], row["backend"])
            if col is None: continue
            try:
                ms = float(row["median_ms"])
            except ValueError:
                continue
            out[(q, col)] = ms
    return out


def fill_summary_from_logs(summary, log_dir):
    """For any (Q, xlsx_col) we DIDN'T get from SUMMARY.csv (shutdown
    crashes truncated the RESULTS summary block), pull the TOTAL
    median from per-iter lines via parse_log_phases().
    """
    filled = 0
    for log_path in sorted(log_dir.glob("q*_*.log")):
        m = re.match(r"q(\d+)_(day|ge)_([^.]+)\.log", log_path.name)
        if not m: continue
        q, pass_, bm = int(m.group(1)), m.group(2), m.group(3)
        if q not in QS_TO_UPDATE: continue
        xlsx_col = col_for(pass_, bm)
        if xlsx_col is None: continue
        if (q, xlsx_col) in summary: continue   # already have it

        label = Q15_SHORT_LABEL.get(bm, "") if q == 15 else BACKEND_LABEL_FOR.get(bm, "")
        phases = parse_log_phases(log_path, label)
        if "TOTAL" in phases:
            summary[(q, xlsx_col)] = phases["TOTAL"]
            filled += 1
    return filled


def parse_log_memory(log_path):
    """
    {bitmap_col: {'total_mb': float, 'breakdown': str|None,
                  'backend_name': str}}
    parsed from one log file.
    """
    text = log_path.read_text()
    result = {}
    # [load_bitmap] X: built BACKEND_NAME index (N keys, X.XXX MB) in ...
    for m in re.finditer(
        r"\[load_bitmap\]\s+(\S+):\s+built\s+(\w+)\s+index\s+\(\d+\s+keys,\s+([\d.]+)\s+MB\)",
        text):
        bm_col, backend_name, mb = m.group(1), m.group(2), float(m.group(3))
        result[bm_col] = {
            "total_mb":     mb,
            "backend_name": backend_name,
            "breakdown":    None,
        }
    # [load_bitmap_breakdown] X BACKEND_NAME: ultra=A MB L1=B MB ...
    for m in re.finditer(
        r"\[load_bitmap_breakdown\]\s+(\S+)\s+(\w+):\s+(.+)$",
        text, re.MULTILINE):
        bm_col, _, body = m.group(1), m.group(2), m.group(3)
        if bm_col in result:
            result[bm_col]["breakdown"] = body.strip()
    return result


def parse_log_phases(log_path, backend_label):
    """
    Return {phase_tag: median_ms} parsed from the per-iter lines of
    one log file.  The first iter is treated as warm-up and dropped
    (matches DEBIT_WARMUP=1 in the sweep harness).  Phase tags emitted:

      - "PhaseA", "PhaseB", ... (per-iter "PhaseX(label)=N" form)
      - "PhaseA_OR", "PhaseB_Agg" (Q15 per-backend "OR=N Agg=N" form)
      - "TOTAL" (all queries print "Total=N")

    Using per-iter values (rather than the summary block) keeps the
    parser working even when the process crashed during shutdown
    after the last iter — e.g. Q19 cb_ge SIGSEGV.  Mapping these
    tags into the xlsx row labels happens in aggregate_phases_per_q.
    """
    text = log_path.read_text()
    per_iter = defaultdict(list)
    for line in text.splitlines():
        if "Total=" not in line: continue
        # TOTAL
        m = re.search(r"Total=\s*([\d.]+)", line)
        if m: per_iter["TOTAL"].append(float(m.group(1)))
        # "PhaseA(...)=N" pattern (the trailing ')' is optional for safety).
        for pm in re.finditer(r"Phase([A-Z])(?:\s*\([^)]*\))?\s*=\s*([\d.]+)", line):
            per_iter["Phase" + pm.group(1)].append(float(pm.group(2)))
        # Q15 "OR=N Agg=N" pattern (only present on Q15 per-iter lines).
        for pm in re.finditer(r"\b(OR|Agg)=\s*([\d.]+)", line):
            key = "PhaseA_OR" if pm.group(1) == "OR" else "PhaseB_Agg"
            per_iter[key].append(float(pm.group(2)))

    def med(xs):
        if not xs: return None
        ys = sorted(xs)
        n = len(ys)
        return ys[n//2] if n % 2 else (ys[n//2 - 1] + ys[n//2]) / 2

    out = {}
    for k, vs in per_iter.items():
        if len(vs) > 1: vs = vs[1:]   # skip warmup
        v = med(vs)
        if v is not None: out[k] = v
    return out


def parse_log_rows(log_path):
    """Extract result row-count(s) from a log.  Q15 prints 'ties=N',
    Q1 prints 'X (rf,ls) groups', other Qs print 'rows=N'."""
    text = log_path.read_text()
    m = re.search(r"rows=(\d+)", text)
    if m: return int(m.group(1))
    m = re.search(r"\((\d+)\s+(?:rf,ls\s+)?groups\)", text)
    if m: return int(m.group(1))
    m = re.search(r"\((\d+)\s+row\(s\)", text)
    if m: return int(m.group(1))
    return None


# --- Helpers for finding cells by row Q label and column header -----

def header_to_col(ws, header_text, header_row=1):
    """
    Return 1-based column index whose header matches header_text.

    Handles two header conventions:
      - bare:  'CB'   (Time Matrix, Storage Total, Phase Time Detail)
      - dashed: 'CB — DDC' (Memory Breakdown — short tag, em-dash, description)
    """
    target = header_text.strip()
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None: continue
        s = str(v).strip()
        if s == target:
            return col
        # Em-dash form: 'SHORT — DESCRIPTION'
        if s.startswith(target + " —") or s.startswith(target + " --"):
            return col
    return None


def q_label_to_row(ws, q_label, q_col=1, header_row=1):
    """Return 1-based row index where col q_col == q_label."""
    for row in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=row, column=q_col).value
        if v is not None and str(v).strip() == q_label:
            return row
    return None


# --- Bitmap memory aggregation across all logs ----------------------

def aggregate_memory(log_dir):
    """
    Returns:
      bm_mem[(bitmap_col, backend_name)] = {'total_mb': MB, 'breakdown': str}
        ↳ same bitmap loaded under multiple Qs ⇒ stable MB; we record the
          first seen value and ignore the rest.
      q_mem[(Q, xlsx_col)] = total MB summed across that Q's loaded
                              bitmaps for that backend.
    """
    bm_mem = {}
    q_mem  = {}
    for log_path in sorted(log_dir.glob("q*_*.log")):
        m = re.match(r"q(\d+)_(day|ge)_([^.]+)\.log", log_path.name)
        if not m: continue
        q, pass_, bm = int(m.group(1)), m.group(2), m.group(3)
        if q not in QS_TO_UPDATE: continue
        xlsx_col = col_for(pass_, bm)
        if xlsx_col is None:
            continue

        memdata = parse_log_memory(log_path)
        q_total = 0.0
        for bm_col, d in memdata.items():
            key = (bm_col, d["backend_name"])
            if key not in bm_mem:
                bm_mem[key] = {"total_mb": d["total_mb"],
                               "breakdown": d["breakdown"]}
            q_total += d["total_mb"]
        if q_total > 0:
            q_mem[(q, xlsx_col)] = q_total
    return bm_mem, q_mem


# --- Bitmaps_for_TPCH_SF10.xlsx update -----------------------------

# Bitmap row label as it appears in the Bitmaps_for_TPCH_SF10 sheet,
# keyed by the bitmap column name in our [load_bitmap] log lines.
BITMAPS_ROW_LABEL = {
    "shipdate":          "shipdate",
    "shipdate_GE_month": "shipdate_GE_month",   # may need to add a new row
    "shipdate_GE":       "shipdate_GE",
    "linestatus":        "linestatus",
    "returnflag":        "returnflag",
    "orderkey":          "orderkey",
    "suppkey":           "suppkey",
    "partkey":           "partkey",
    "discount":          "discount",
    "quantity":          "quantity",
    "shipmode":          "shipmode",
    "shipinstruct":      "shipinstruct",
    "receiptdate":       "receiptdate",
    "receiptdate_GE":    "receiptdate_GE",
}

# Backend column header in Bitmaps_for_TPCH_SF10 sheet.
BITMAPS_COL_FOR_BACKEND = {
    "WAH":         "WAH",
    "DDC":      "DDC",
    "DDCGE":    "DDC",      # same column — GE variant is built from same algorithm
    "DDCBPE":   "DDC+BPE",
    "CRoaring":    "CRoaring",
    "CRoaringBPE": "CRoaring+BPE",
    "CRoaringRun": "CRoaring+Run",
    "EWAH":        "EWAH",
    "Concise":     "Concise",
}


def update_bitmaps_xlsx(path, bm_mem):
    wb = xl.load_workbook(path)
    ws = wb.active

    updated = []
    added   = []

    # Header row is row 2 (row 1 is the spreadsheet title).
    HEADER_ROW = 2

    # Locate the Total row up-front — new bitmap rows go IMMEDIATELY
    # above it so the data block stays contiguous and the Total row
    # always sums all real bitmap rows (incl. newly added ones).
    def find_total_row():
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).strip().lower() == "total":
                return r
        return None

    total_row = find_total_row()

    # Map row_label -> row idx for fast existing-row lookup.
    label_to_row = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            label_to_row[str(v).strip()] = r

    new_row_for = {}   # row_label -> row idx (for newly-added rows)
    for (bm_col, backend_name), d in bm_mem.items():
        row_label = BITMAPS_ROW_LABEL.get(bm_col, bm_col)
        col_label = BITMAPS_COL_FOR_BACKEND.get(backend_name)
        if col_label is None:
            continue  # Bitset etc. — not in our sweep

        row = label_to_row.get(row_label) or new_row_for.get(row_label)
        if row is None:
            # Insert a new row immediately above the Total row.
            assert total_row is not None, "Total row missing from Bitmaps sheet"
            ws.insert_rows(total_row)
            row = total_row
            ws.cell(row=row, column=1).value = row_label
            ws.cell(row=row, column=2).value = None     # cardinality unknown
            ws.cell(row=row, column=3).value = "Added by post-fix sweep (auto)"
            new_row_for[row_label] = row
            added.append(row_label)
            # Total row + all rows below shift down by 1.
            total_row += 1
            # Shift any cached new-row indices that were above the insertion point
            # (none, since we always insert at total_row).  Existing rows above
            # the insertion point keep their indices.

        col = header_to_col(ws, col_label, header_row=HEADER_ROW)
        if col is None:
            continue

        old = ws.cell(row=row, column=col).value
        new = round(d["total_mb"], 4)
        ws.cell(row=row, column=col).value = new
        updated.append((row_label, col_label, old, new))

    # Recompute Total = sum of all numeric values in this column from
    # the data rows (everything between HEADER_ROW+1 and total_row,
    # excluding total_row itself).
    if total_row is not None:
        for c in range(4, ws.max_column + 1):
            header = ws.cell(row=HEADER_ROW, column=c).value
            if header is None: continue
            s = 0.0
            any_filled = False
            for r in range(HEADER_ROW + 1, total_row):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)):
                    s += float(v); any_filled = True
            ws.cell(row=total_row, column=c).value = round(s, 2) if any_filled else None

    wb.save(path)
    return updated, added


# --- bm_results.xlsx update ----------------------------------------

def update_time_matrix(ws, summary):
    """summary: {(Q, col): ms}.  Update only Qs in QS_TO_UPDATE.
    Also mirrors CB→CB+BPE and CR→CR+BPE for Qs in BPE_MIRRORS_NONBPE
    (where the BPE column is algorithmically identical to non-BPE)."""
    log = []
    for (q, col), ms in summary.items():
        q_label = f"Q{q}"
        row = q_label_to_row(ws, q_label)
        cidx = header_to_col(ws, col)
        if row and cidx:
            old = ws.cell(row=row, column=cidx).value
            ws.cell(row=row, column=cidx).value = round(ms, 4)
            log.append((q_label, col, old, round(ms, 4)))

    # BPE mirroring pass.
    for q in BPE_MIRRORS_NONBPE & QS_TO_UPDATE:
        q_label = f"Q{q}"
        row = q_label_to_row(ws, q_label)
        if row is None: continue
        for src, dst in (("CB", "CB+BPE"), ("CR", "CR+BPE")):
            sc = header_to_col(ws, src)
            dc = header_to_col(ws, dst)
            if sc and dc:
                v = ws.cell(row=row, column=sc).value
                if v is not None:
                    old = ws.cell(row=row, column=dc).value
                    ws.cell(row=row, column=dc).value = v
                    log.append((q_label, f"{dst}←{src}", old, v))
    return log


def update_phase_detail(ws, phases_data):
    """
    phases_data: {(Q, xlsx_col): {phase_name: ms}}
    The Phase Time Detail sheet has a Q label in col 1 on the first
    row of each Q's block, blank for subsequent phases in the same
    block.  We index by phase_name in col 2.
    """
    log = []

    # Walk the sheet and build {(Q, phase): row_idx}.
    row_for = {}
    cur_q = None
    for r in range(2, ws.max_row + 1):
        q_cell = ws.cell(row=r, column=1).value
        p_cell = ws.cell(row=r, column=2).value
        if q_cell:
            cur_q = str(q_cell).strip()
        if p_cell and cur_q:
            row_for[(cur_q, str(p_cell).strip())] = r

    for (q, col), phases in phases_data.items():
        q_label = f"Q{q}"
        cidx = header_to_col(ws, col)
        if cidx is None: continue
        for phase, ms in phases.items():
            r = row_for.get((q_label, phase))
            if r is None:
                continue
            old = ws.cell(row=r, column=cidx).value
            ws.cell(row=r, column=cidx).value = round(ms, 4)
            log.append((q_label, phase, col, old, round(ms, 4)))
    return log


def update_storage_total(ws, q_mem):
    log = []
    for (q, col), mb in q_mem.items():
        q_label = f"Q{q}"
        r = q_label_to_row(ws, q_label)
        c = header_to_col(ws, col)
        if r and c:
            old = ws.cell(row=r, column=c).value
            ws.cell(row=r, column=c).value = round(mb, 4)
            log.append((q_label, col, old, round(mb, 4)))
    return log


def update_memory_breakdown(ws, q_breakdown):
    """
    q_breakdown: {(Q, xlsx_col): str} — multi-line breakdown text
    Format mirrors existing cells: one line per category, joined with \n.
    """
    log = []
    for (q, col), text in q_breakdown.items():
        q_label = f"Q{q}"
        r = q_label_to_row(ws, q_label)
        c = header_to_col(ws, col)
        if r and c:
            ws.cell(row=r, column=c).value = text
            log.append((q_label, col))
    return log


def mirror_bpe_columns(ws, q_label_col=1, header_row=1):
    """For each Q in BPE_MIRRORS_NONBPE & QS_TO_UPDATE, copy the CB
    column value to CB+BPE and CR to CR+BPE.  Works on Time Matrix,
    Storage Total (MB), Memory Breakdown, and Correctness sheets.
    """
    n = 0
    for q in BPE_MIRRORS_NONBPE & QS_TO_UPDATE:
        q_label = f"Q{q}"
        row = q_label_to_row(ws, q_label, q_col=q_label_col,
                             header_row=header_row)
        if row is None: continue
        for src, dst in (("CB", "CB+BPE"), ("CR", "CR+BPE")):
            sc = header_to_col(ws, src, header_row=header_row)
            dc = header_to_col(ws, dst, header_row=header_row)
            # Correctness has 'X rows' headers — try that form too.
            if sc is None: sc = header_to_col(ws, src + " rows", header_row=header_row)
            if dc is None: dc = header_to_col(ws, dst + " rows", header_row=header_row)
            if sc and dc:
                v = ws.cell(row=row, column=sc).value
                if v is not None:
                    ws.cell(row=row, column=dc).value = v
                    n += 1
    return n


def mirror_bpe_phase_detail(ws):
    """Phase Time Detail keeps a multi-row block per Q (PhaseA, PhaseB,
    …, TOTAL).  Mirror every phase row's CB column into CB+BPE (same
    for CR/CR+BPE) for Qs in BPE_MIRRORS_NONBPE."""
    n = 0
    # Build {(Q_label, phase): row}
    row_for = {}
    cur_q = None
    for r in range(2, ws.max_row + 1):
        q_cell = ws.cell(row=r, column=1).value
        p_cell = ws.cell(row=r, column=2).value
        if q_cell: cur_q = str(q_cell).strip()
        if p_cell and cur_q:
            row_for[(cur_q, str(p_cell).strip())] = r

    for q in BPE_MIRRORS_NONBPE & QS_TO_UPDATE:
        q_label = f"Q{q}"
        for (ql, ph), r in row_for.items():
            if ql != q_label: continue
            for src, dst in (("CB", "CB+BPE"), ("CR", "CR+BPE")):
                sc = header_to_col(ws, src)
                dc = header_to_col(ws, dst)
                if sc and dc:
                    v = ws.cell(row=r, column=sc).value
                    if v is not None:
                        ws.cell(row=r, column=dc).value = v
                        n += 1
    return n


def update_correctness(ws, q_rows):
    """q_rows: {(Q, col_header_for_correctness): row_count}.
    Correctness has '<COL> rows' headers."""
    log = []
    for (q, col_simple), cnt in q_rows.items():
        q_label = f"Q{q}"
        r = q_label_to_row(ws, q_label)
        c = header_to_col(ws, col_simple + " rows")
        if r and c:
            old = ws.cell(row=r, column=c).value
            ws.cell(row=r, column=c).value = cnt
            log.append((q_label, col_simple, old, cnt))
    return log


# --- Phase / breakdown extraction for all (Q, col) combinations ----

# Per-Q mapping from short phase tag in the log ("PhaseA", "PhaseD", …)
# to the row label used by the Phase Time Detail sheet.  These must
# stay in lock-step with the xlsx schema — phases not in the map are
# silently dropped (Q15's OR/Agg use the Q15-specific labels which
# match the xlsx already).
PHASE_LABEL = {
    1:  {"PhaseA": "PhaseA", "PhaseB": "PhaseB",
         "PhaseC": "PhaseC", "TOTAL": "TOTAL"},
    10: {"PhaseA": "PhaseA_orders", "PhaseB": "PhaseB_okey",
         "PhaseC": "PhaseC_rf",     "PhaseD": "PhaseD_AND",
         "PhaseE": "PhaseE_agg",    "PhaseF": "PhaseF_top20",
         "TOTAL":  "TOTAL"},
    14: {"PhaseA": "PhaseA_ship", "PhaseB": "PhaseB_part",
         "PhaseC": "PhaseC_agg",  "TOTAL":  "TOTAL"},
    15: {"PhaseA_OR": "PhaseA_OR", "PhaseB_Agg": "PhaseB_Agg",
         "TOTAL":     "TOTAL"},
    17: {"PhaseA": "PhaseA_part", "PhaseB": "PhaseB_avg",
         "PhaseC": "PhaseC_sum",  "TOTAL":  "TOTAL"},
    19: {"PhaseA": "PhaseA_ship", "PhaseB": "PhaseB_part",
         "PhaseC": "PhaseC_agg",  "TOTAL":  "TOTAL"},
}


def aggregate_phases_per_q(log_dir):
    """
    {(Q, xlsx_col): {xlsx_phase_label: ms}}
    """
    out = defaultdict(dict)
    for log_path in sorted(log_dir.glob("q*_*.log")):
        m = re.match(r"q(\d+)_(day|ge)_([^.]+)\.log", log_path.name)
        if not m: continue
        q, pass_, bm = int(m.group(1)), m.group(2), m.group(3)
        if q not in QS_TO_UPDATE: continue
        xlsx_col = col_for(pass_, bm)
        if xlsx_col is None: continue

        label = Q15_SHORT_LABEL.get(bm, "") if q == 15 else BACKEND_LABEL_FOR.get(bm, "")
        raw_phases = parse_log_phases(log_path, label)
        mp = PHASE_LABEL.get(q, {})
        mapped = {}
        for k, v in raw_phases.items():
            xlsx_label = mp.get(k)
            if xlsx_label is not None:
                mapped[xlsx_label] = v
        if mapped:
            out[(q, xlsx_col)] = mapped
    return out


def aggregate_breakdown_per_q(log_dir):
    """
    {(Q, xlsx_col): str_multiline}
    """
    by_q_col = defaultdict(list)   # collect (bitmap, line)
    for log_path in sorted(log_dir.glob("q*_*.log")):
        m = re.match(r"q(\d+)_(day|ge)_([^.]+)\.log", log_path.name)
        if not m: continue
        q, pass_, bm = int(m.group(1)), m.group(2), m.group(3)
        if q not in QS_TO_UPDATE: continue
        xlsx_col = col_for(pass_, bm)
        if xlsx_col is None: continue
        memdata = parse_log_memory(log_path)
        # Build a per-line breakdown for the loaded bitmaps in this Q.
        # Format: per-category line per bitmap, like the existing xlsx:
        #   "ultra: 0.00 MB\nL1: 89.33 MB\nL2: 61.99 MB\n..."
        # But the original aggregates all bitmaps under the Q's totals.
        # We follow the existing format: one line per "<cat>: X.XX MB"
        # listed in the order they appear, summed across loaded bitmaps.
        cat_sum = defaultdict(float)
        for bm_col, d in memdata.items():
            if d["breakdown"] is None: continue
            # body example: "ultra=0 MB L1=89.3341 MB L2=61.98755 MB ..."
            for cm in re.finditer(r"(\w+)=([\d.]+)\s*MB", d["breakdown"]):
                cat_sum[cm.group(1)] += float(cm.group(2))
        if cat_sum:
            lines = [f"{k}: {v:.2f} MB" for k, v in cat_sum.items()]
            by_q_col[(q, xlsx_col)] = "\n".join(lines)
    return by_q_col


def aggregate_rows_per_q(log_dir):
    """{(Q, simple_col): row_count}, simple_col = 'WAH'/'CB'/etc.
    For Correctness sheet headers like 'WAH rows'."""
    out = {}
    for log_path in sorted(log_dir.glob("q*_*.log")):
        m = re.match(r"q(\d+)_(day|ge)_([^.]+)\.log", log_path.name)
        if not m: continue
        q, pass_, bm = int(m.group(1)), m.group(2), m.group(3)
        if q not in QS_TO_UPDATE: continue
        xlsx_col = col_for(pass_, bm)
        if xlsx_col is None: continue
        n = parse_log_rows(log_path)
        if n is not None:
            out[(q, xlsx_col)] = n
    return out


# --- Memory Detail (long format) helper -----------------------------

def update_memory_detail(ws, log_dir):
    """
    Memory Detail rows: Q | Backend | Backend (full) | Category | MB.
    Re-write rows for QS_TO_UPDATE; preserve other Qs.
    Existing rows for QS_TO_UPDATE are removed and replaced.
    """
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # Snapshot rows we keep (those whose Q is NOT in QS_TO_UPDATE).
    keep = []
    for r in range(2, ws.max_row + 1):
        q_val = ws.cell(row=r, column=1).value
        if q_val is None: continue
        m = re.match(r"Q(\d+)", str(q_val))
        if not m: continue
        if int(m.group(1)) in QS_TO_UPDATE: continue
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        keep.append(row_vals)

    # New rows from logs.
    new_rows = []
    cb_rows_by_q = defaultdict(list)   # for BPE mirror
    cr_rows_by_q = defaultdict(list)
    for log_path in sorted(log_dir.glob("q*_*.log")):
        m = re.match(r"q(\d+)_(day|ge)_([^.]+)\.log", log_path.name)
        if not m: continue
        q, pass_, bm = int(m.group(1)), m.group(2), m.group(3)
        if q not in QS_TO_UPDATE: continue
        xlsx_col = col_for(pass_, bm)
        if xlsx_col is None: continue

        cat_sum = defaultdict(float)
        backend_full = ""
        for bm_col, mem in parse_log_memory(log_path).items():
            backend_full = mem["backend_name"]
            if mem["breakdown"]:
                for cm in re.finditer(r"(\w+)=([\d.]+)\s*MB", mem["breakdown"]):
                    cat_sum[cm.group(1)] += float(cm.group(2))
            cat_sum["total"] += mem["total_mb"]
        for cat, mb in cat_sum.items():
            row_data = [f"Q{q}", xlsx_col, backend_full, cat, mb]
            new_rows.append(row_data)
            if xlsx_col == "CB":  cb_rows_by_q[q].append(row_data)
            if xlsx_col == "CR":  cr_rows_by_q[q].append(row_data)

    # BPE-mirror pass: for Qs where BPE is algorithmically identical to
    # non-BPE, duplicate the CB / CR rows as CB+BPE / CR+BPE rows.
    for q in BPE_MIRRORS_NONBPE & QS_TO_UPDATE:
        for src_row in cb_rows_by_q.get(q, []):
            new_rows.append([src_row[0], "CB+BPE",
                             "DDC + BPE (β scheme)",
                             src_row[3], src_row[4]])
        for src_row in cr_rows_by_q.get(q, []):
            new_rows.append([src_row[0], "CR+BPE",
                             "CRoaring + BPE (β scheme)",
                             src_row[3], src_row[4]])

    # Clear sheet rows below header.
    for r in range(ws.max_row, 1, -1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    # Write keep + new sorted by Q then Backend then Category.
    all_rows = keep + new_rows
    all_rows.sort(key=lambda row: (
        int(re.match(r"Q(\d+)", str(row[0])).group(1)) if re.match(r"Q(\d+)", str(row[0])) else 9999,
        str(row[1]),
        str(row[3]),
    ))
    for i, row in enumerate(all_rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c).value = v


# --- Main -----------------------------------------------------------

def main():
    if not (LOG_DIR / "SUMMARY.csv").exists():
        sys.exit(f"missing {LOG_DIR/'SUMMARY.csv'} — run the sweep first")

    print(f"[parse]  SUMMARY.csv ...")
    summary = parse_summary_csv(LOG_DIR / "SUMMARY.csv")
    print(f"         {len(summary)} (Q,col) total-ms entries from CSV")
    n_filled = fill_summary_from_logs(summary, LOG_DIR)
    if n_filled:
        print(f"         +{n_filled} filled from per-iter log Total= (shutdown-truncated logs)")

    print(f"[parse]  per-bitmap memory ...")
    bm_mem, q_mem = aggregate_memory(LOG_DIR)
    print(f"         {len(bm_mem)} (bm,backend) entries; {len(q_mem)} (Q,col) per-Q totals")

    print(f"[parse]  phase breakdowns ...")
    phases = aggregate_phases_per_q(LOG_DIR)
    print(f"         {len(phases)} (Q,col) phase entries")

    print(f"[parse]  per-Q breakdown strings ...")
    breakdown = aggregate_breakdown_per_q(LOG_DIR)
    print(f"         {len(breakdown)} (Q,col) breakdown entries")

    print(f"[parse]  row-counts ...")
    rows_cnt = aggregate_rows_per_q(LOG_DIR)
    print(f"         {len(rows_cnt)} (Q,col) row entries")

    # ----- Bitmaps_for_TPCH_SF10.xlsx -----
    print(f"\n[write]  {BITMAPS_XL.name}")
    updated, added = update_bitmaps_xlsx(BITMAPS_XL, bm_mem)
    print(f"         updated {len(updated)} cells, added {len(added)} rows")
    if added:
        print(f"         new rows: {added}")
    for r, c, old, new in updated[:8]:
        print(f"           {r:18s} / {c:14s} : {old} → {new}")
    if len(updated) > 8: print(f"           ... +{len(updated)-8} more")

    # ----- bm_results.xlsx -----
    print(f"\n[write]  {RESULTS_XL.name}")
    wb = xl.load_workbook(RESULTS_XL)

    print(f"  Time Matrix:")
    tm = update_time_matrix(wb["Time Matrix"], summary)
    print(f"           {len(tm)} cells (incl. BPE mirroring)")

    print(f"  Phase Time Detail:")
    pd_ = update_phase_detail(wb["Phase Time Detail"], phases)
    n_bpe = mirror_bpe_phase_detail(wb["Phase Time Detail"])
    print(f"           {len(pd_)} cells written; {n_bpe} cells mirrored CB→CB+BPE/CR→CR+BPE")

    print(f"  Storage Total (MB):")
    st = update_storage_total(wb["Storage Total (MB)"], q_mem)
    n_bpe = mirror_bpe_columns(wb["Storage Total (MB)"])
    print(f"           {len(st)} cells written; {n_bpe} BPE mirrored")

    print(f"  Memory Breakdown:")
    mb_ = update_memory_breakdown(wb["Memory Breakdown"], breakdown)
    n_bpe = mirror_bpe_columns(wb["Memory Breakdown"])
    print(f"           {len(mb_)} cells written; {n_bpe} BPE mirrored")

    print(f"  Memory Detail (MB):")
    update_memory_detail(wb["Memory Detail (MB)"], LOG_DIR)
    print(f"           regenerated (incl. BPE rows mirrored from CB / CR)")

    print(f"  Correctness:")
    cr_ = update_correctness(wb["Correctness"], rows_cnt)
    n_bpe = mirror_bpe_columns(wb["Correctness"])
    print(f"           {len(cr_)} cells written; {n_bpe} BPE mirrored")

    wb.save(RESULTS_XL)
    print(f"\n[done]  saved {RESULTS_XL.name}")


if __name__ == "__main__":
    main()
