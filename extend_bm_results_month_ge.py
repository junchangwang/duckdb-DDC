#!/usr/bin/env python3
"""Update bm_results.xlsx with CB+GE / CR+GE columns from the month-GE
sweep logs (bm_logs_month_ge_cb_ge/ and bm_logs_month_ge_cr/).

This is a re-runnable updater:
  - Backs up the existing xlsx (timestamped, into bm_result_backups/).
  - If CB+GE / CR+GE columns already exist, they are removed and re-added
    so the data reflects the latest sweep.
  - Parses per-phase median (across non-warmup iterations) and fills the
    matching rows in Phase Time Detail.
  - Adds 3-color-scale conditional formatting (green→yellow→red, percentile
    50) so the new columns are coloured the same as the existing ones.

CB+GE := DEBIT_BM=cb_ge sweep — month-GE for Q1/Q3/Q14/Q15, year-GE for Q6.
CR+GE := DEBIT_BM=cr   sweep — same column layout (CR builds month-GE too).
"""

from __future__ import annotations

import re
import shutil
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule, FormatObject, ColorScale
from openpyxl.styles import Alignment, Color
from openpyxl.utils import get_column_letter, column_index_from_string

REPO = Path(__file__).resolve().parent
XLSX = REPO / "bm_results.xlsx"
LOG_CB = REPO / "bm_logs_month_ge_cb_ge"
LOG_CR = REPO / "bm_logs_month_ge_cr"

QUERIES = [1, 3, 4, 5, 6, 8, 10, 12, 14, 15, 17, 19]

NEW_BACKENDS: List[Tuple[str, str, List[str]]] = [
    ("CB+GE", "DDC + GE (month/year)", ["ultra", "L1", "L2", "L3", "L4", "header"]),
    ("CR+GE", "CRoaring + GE (month/year)", ["array", "bitset", "run", "header"]),
]
NEW_BACKEND_NAMES = [bn for bn, _, _ in NEW_BACKENDS]
NEW_BACKEND_FULLNAMES = {bn: full for bn, full, _ in NEW_BACKENDS}
NEW_BACKEND_LAYERS = {bn: lay for bn, _, lay in NEW_BACKENDS}

# -- regex ----------------------------------------------------------------

RE_BUILT = re.compile(
    r"^\[load_bitmap\](?:\s+\(CB_GE auto\))?\s+(\S+):\s+built\s+(\w+)\s+index\s+"
    r"\(\d+\s+keys,\s+([\d.]+)\s+MB\)"
)
RE_BREAKDOWN = re.compile(
    r"^\[load_bitmap_breakdown\](?:\s+\(CB_GE auto\))?\s+(\S+)\s+(\w+):\s+(.+)$"
)
RE_KV = re.compile(r"(\w+)=([\d.]+)\s*MB")
RE_TOTAL = re.compile(r"TOTAL\s*:\s*([\d.]+)")
RE_LOWER_TOTAL = re.compile(r"[tT]otal=([\d.]+)")
RE_OK = re.compile(r"^\[OK\][^\n]*", re.MULTILINE)
RE_FAIL = re.compile(r"^\[FAIL\][^\n]*", re.MULTILINE)
RE_ROWS = re.compile(r"rows[=:]\s*(\d+)")

# Per-iter line: "  <Backend>:  PhaseA(...)=X PhaseB(...)=Y ... Total=Z ..."
# or "  [Q6 DDC] ship_ge=X OR_discount=Y ... total=Z rows=..."
RE_ITER_LINE = re.compile(
    r"^\s+(?:\[Q\d+\s+)?(?:DDC(?:GE)?|CRoaring(?:Run)?|WAH|EWAH|Concise|"
    r"CB(?:_GE|_BPE)?|CR(?:R|_BPE)?|EW|CON):?\s",
)
RE_KV_NUMERIC = re.compile(r"([A-Za-z_][\w()+\s/]*?)=([\d.]+(?:[eE][+\-]?\d+)?)")

PHASE_DROP_TOKENS = {
    "total", "max", "ties", "rows", "revenue", "duckdb",
    "share_1995", "share_1996", "share",
    "promo_pct", "promo_revenue", "rev",
}


def _looks_like_iter_line(line: str) -> bool:
    if not line.startswith(" "):
        return False
    s = line.strip()
    return ("=" in s) and any(
        s.split(":")[0].strip() == k or s.startswith(f"[{k}")
        for k in ("DDC", "DDCGE", "CRoaring", "CRoaringRun", "WAH", "EWAH",
                  "Concise", "CB", "CR", "CRR", "EW", "BS", "BSA", "CON", "[Q6")
    )


def parse_iter_line(line: str) -> List[Tuple[str, float]]:
    """Extract (phase_label, ms) pairs from a per-iter line.  Drops
    Total/rows/max/ties/share/revenue.  Preserves order so caller can
    align against the xlsx phase rows by position."""
    out: List[Tuple[str, float]] = []
    for m in RE_KV_NUMERIC.finditer(line):
        label = m.group(1).strip()
        # Strip "(descr)" suffix to keep just "PhaseA" / "ship_ge" / "OR".
        bare = re.sub(r"\(.*?\)", "", label).strip()
        if bare.lower() in PHASE_DROP_TOKENS:
            continue
        # Filter out scientific-notation max/etc that survived above.
        try:
            val = float(m.group(2))
        except ValueError:
            continue
        out.append((bare, val))
    return out


def parse_log(path: Path) -> Dict:
    text = path.read_text(errors="replace")

    cols: List[Tuple[str, float]] = []
    layers: Dict[str, float] = defaultdict(float)
    for line in text.splitlines():
        m = RE_BUILT.match(line)
        if m:
            col, _bname, mb = m.group(1), m.group(2), float(m.group(3))
            cols.append((col, mb))
            continue
        m = RE_BREAKDOWN.match(line)
        if m:
            for cat, mb in ((k, float(v)) for k, v in RE_KV.findall(m.group(3))):
                layers[cat] += mb

    ms = None
    for m in RE_TOTAL.finditer(text):
        ms = float(m.group(1))
    if ms is None:
        for m in RE_LOWER_TOTAL.finditer(text):
            ms = float(m.group(1))

    # Per-iter phase values (positional list per iter).  Skip the first
    # iteration (warmup).  Take median per position across iters.
    iter_phases: List[List[Tuple[str, float]]] = []
    for line in text.splitlines():
        if not _looks_like_iter_line(line):
            continue
        kvs = parse_iter_line(line)
        if kvs:
            iter_phases.append(kvs)
    phase_median: List[Tuple[str, float]] = []
    if len(iter_phases) > 1:
        measured = iter_phases[1:]   # drop warmup
        # Use the first measured iter's labels as the order template.
        labels = [lbl for lbl, _ in measured[0]]
        for i, lbl in enumerate(labels):
            vals: List[float] = []
            for it in measured:
                if i < len(it) and it[i][0] == lbl:
                    vals.append(it[i][1])
            if vals:
                phase_median.append((lbl, statistics.median(vals)))

    ok = RE_OK.findall(text)
    fail = RE_FAIL.findall(text)
    if fail:
        status, msg = "FAIL", fail[-1]
    elif ok:
        status, msg = "OK", ok[-1]
    else:
        status, msg = "?", ""
    rows = None
    rrows = RE_ROWS.findall(text)
    if rrows:
        rows = int(rrows[-1])

    return {
        "ms": ms,
        "status": status,
        "msg": msg,
        "rows": rows,
        "cols": cols,
        "layers": dict(layers),
        "phase_median": phase_median,
    }


def load_all(log_dir: Path) -> Dict[int, Dict]:
    out = {}
    for q in QUERIES:
        log = log_dir / f"q{q}.log"
        if log.exists():
            out[q] = parse_log(log)
    return out


def render_memory_cell(rec: Dict, layer_order: List[str]) -> str:
    lines = []
    for cat in layer_order:
        lines.append(f"{cat}: {rec['layers'].get(cat, 0.0):.2f} MB")
    for name, mb in rec["cols"]:
        lines.append(f"{name}: {mb:.2f} MB")
    grand = sum(mb for _, mb in rec["cols"])
    lines.append(f"total: {grand:.2f} MB")
    return "\n".join(lines)


# -- xlsx helpers ----------------------------------------------------------

def find_q_row(ws, q: int) -> int | None:
    target = f"Q{q}"
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == target:
            return row
    return None


def find_existing_columns(ws, names: List[str]) -> Dict[str, int]:
    """Return 1-indexed col positions for any header that mentions one of the names."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if not v:
            continue
        for name in names:
            if name in str(v):
                out[name] = c
    return out


def delete_columns(ws, cols_to_remove: List[int]) -> None:
    """Delete by index, descending so earlier indices stay stable."""
    for c in sorted(cols_to_remove, reverse=True):
        ws.delete_cols(c, 1)


def make_color_scale_rule() -> ColorScaleRule:
    """3-stop green→yellow→red, percentile 50 — matches existing rules."""
    return ColorScaleRule(
        start_type="min",  start_color="63BE7B",   # green
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
        end_type="max",    end_color="F8696B",     # red
    )


def add_per_row_color_scale(ws, rows: List[int], start_col: int, end_col: int) -> None:
    """Apply a 3-stop color scale on each row over [start_col..end_col]."""
    for r in rows:
        a = f"{get_column_letter(start_col)}{r}:{get_column_letter(end_col)}{r}"
        ws.conditional_formatting.add(a, make_color_scale_rule())


def add_per_column_color_scale(ws, cols: List[int], start_row: int, end_row: int) -> None:
    """Apply on each column over [start_row..end_row]."""
    for c in cols:
        letter = get_column_letter(c)
        a = f"{letter}{start_row}:{letter}{end_row}"
        ws.conditional_formatting.add(a, make_color_scale_rule())


# -- main ------------------------------------------------------------------

def main():
    if not XLSX.exists():
        raise SystemExit(f"missing: {XLSX}")
    cb = load_all(LOG_CB)
    cr = load_all(LOG_CR)
    new_recs = {"CB+GE": cb, "CR+GE": cr}

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = REPO / "bm_result_backups"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"bm_results_pre_month_ge_v2_{ts}.xlsx"
    shutil.copy2(XLSX, backup)
    print(f"[backup] {XLSX} -> {backup}")

    wb = openpyxl.load_workbook(XLSX)

    # ----- Drop any prior CB+GE / CR+GE columns -----
    for sn in ["Time Matrix", "Phase Time Detail",
               "Memory Breakdown", "Storage Total (MB)", "Correctness"]:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        # Find header strings containing CB+GE or CR+GE.
        existing = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v and any(name in str(v) for name in NEW_BACKEND_NAMES):
                existing.append(c)
        if existing:
            delete_columns(ws, existing)
            print(f"[clean] {sn}: removed {len(existing)} stale CB+GE/CR+GE columns")

    # Memory Detail (MB) — drop old rows for CB+GE / CR+GE
    ws = wb["Memory Detail (MB)"]
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v in NEW_BACKEND_NAMES:
            rows_to_delete.append(r)
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)
    if rows_to_delete:
        print(f"[clean] Memory Detail (MB): removed {len(rows_to_delete)} stale rows")

    # ----- Time Matrix -----
    ws = wb["Time Matrix"]
    base_start, base_end = 2, 11   # B..K = the 10 existing backends
    new_cols_tm = []
    for bn in NEW_BACKEND_NAMES:
        c = ws.max_column + 1
        ws.cell(row=1, column=c).value = bn
        new_cols_tm.append(c)
        for q in QUERIES:
            r = find_q_row(ws, q)
            if r is None:
                continue
            rec = new_recs[bn].get(q)
            if rec and rec["ms"] is not None:
                ws.cell(row=r, column=c).value = rec["ms"]
    # Per-row colour scale over the whole data span B..N (incl. new cols).
    # Each Q-row gets its own scale so coloration is computed across all backends in that row.
    rows = [find_q_row(ws, q) for q in QUERIES if find_q_row(ws, q)]
    add_per_row_color_scale(ws, rows, base_start, new_cols_tm[-1])

    # ----- Phase Time Detail (per-phase median + TOTAL) -----
    ws = wb["Phase Time Detail"]
    new_cols_phase = []
    for bn in NEW_BACKEND_NAMES:
        c = ws.max_column + 1
        ws.cell(row=1, column=c).value = bn
        new_cols_phase.append(c)
        # Walk each Q's phase rows in order and zip with parsed phase medians.
        current_q = None
        q_phase_idx = 0
        for r in range(2, ws.max_row + 1):
            q_cell = ws.cell(row=r, column=1).value
            phase_cell = ws.cell(row=r, column=2).value
            if q_cell and isinstance(q_cell, str) and q_cell.startswith("Q"):
                try:
                    current_q = int(q_cell[1:])
                except ValueError:
                    current_q = None
                q_phase_idx = 0
            if not phase_cell or current_q is None:
                continue
            rec = new_recs[bn].get(current_q)
            if not rec:
                continue
            label = str(phase_cell).strip()
            if label.upper() == "TOTAL":
                if rec["ms"] is not None:
                    ws.cell(row=r, column=c).value = rec["ms"]
                continue
            # Match by position: q_phase_idx-th non-TOTAL phase row.
            pm = rec.get("phase_median", [])
            if q_phase_idx < len(pm):
                ws.cell(row=r, column=c).value = pm[q_phase_idx][1]
            q_phase_idx += 1
    # CF: per row spanning C..(last new col).  C is the first backend col.
    # Find the first numeric backend column (col index of "WAH" header).
    phase_first_col = None
    for cc in range(3, ws.max_column + 1):
        if ws.cell(row=1, column=cc).value == "WAH":
            phase_first_col = cc
            break
    if phase_first_col is None:
        phase_first_col = 3
    phase_last_col = new_cols_phase[-1]
    rows_phase = [r for r in range(2, ws.max_row + 1)
                  if ws.cell(row=r, column=2).value]  # any row with a Phase label
    add_per_row_color_scale(ws, rows_phase, phase_first_col, phase_last_col)

    # ----- Memory Breakdown (text cells) -----
    ws = wb["Memory Breakdown"]
    for bn in NEW_BACKEND_NAMES:
        c = ws.max_column + 1
        ws.cell(row=1, column=c).value = f"{bn} — {NEW_BACKEND_FULLNAMES[bn]}"
        for q in QUERIES:
            r = find_q_row(ws, q)
            if r is None:
                continue
            rec = new_recs[bn].get(q)
            if not rec:
                continue
            cell = ws.cell(row=r, column=c)
            cell.value = render_memory_cell(rec, NEW_BACKEND_LAYERS[bn])
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ----- Memory Detail (MB) — long format -----
    ws = wb["Memory Detail (MB)"]
    start = ws.max_row + 1
    for bn in NEW_BACKEND_NAMES:
        for q in QUERIES:
            rec = new_recs[bn].get(q)
            if not rec:
                continue
            for cat in NEW_BACKEND_LAYERS[bn]:
                ws.cell(row=start, column=1).value = f"Q{q}"
                ws.cell(row=start, column=2).value = bn
                ws.cell(row=start, column=3).value = NEW_BACKEND_FULLNAMES[bn]
                ws.cell(row=start, column=4).value = cat
                ws.cell(row=start, column=5).value = round(rec["layers"].get(cat, 0.0), 4)
                start += 1
            for name, mb in rec["cols"]:
                ws.cell(row=start, column=1).value = f"Q{q}"
                ws.cell(row=start, column=2).value = bn
                ws.cell(row=start, column=3).value = NEW_BACKEND_FULLNAMES[bn]
                ws.cell(row=start, column=4).value = name
                ws.cell(row=start, column=5).value = round(mb, 4)
                start += 1
            total = sum(mb for _, mb in rec["cols"])
            ws.cell(row=start, column=1).value = f"Q{q}"
            ws.cell(row=start, column=2).value = bn
            ws.cell(row=start, column=3).value = NEW_BACKEND_FULLNAMES[bn]
            ws.cell(row=start, column=4).value = "total"
            ws.cell(row=start, column=5).value = round(total, 4)
            start += 1

    # ----- Storage Total (MB) -----
    ws = wb["Storage Total (MB)"]
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        insert_at = hdrs.index("On-disk Σ") + 1
    except ValueError:
        insert_at = ws.max_column + 1
    ws.insert_cols(insert_at, amount=len(NEW_BACKEND_NAMES))
    new_cols_storage = []
    for i, bn in enumerate(NEW_BACKEND_NAMES):
        c = insert_at + i
        ws.cell(row=1, column=c).value = bn
        new_cols_storage.append(c)
        for q in QUERIES:
            r = find_q_row(ws, q)
            if r is None:
                continue
            rec = new_recs[bn].get(q)
            if not rec:
                continue
            total = sum(mb for _, mb in rec["cols"])
            ws.cell(row=r, column=c).value = round(total, 4)
    # CF: per row B..(last new) skipping non-data tail.
    rows_storage = [find_q_row(ws, q) for q in QUERIES if find_q_row(ws, q)]
    storage_last_col = new_cols_storage[-1]
    add_per_row_color_scale(ws, rows_storage, 2, storage_last_col)

    # ----- Correctness — append "rows" cols -----
    ws = wb["Correctness"]
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        insert_at = hdrs.index("DuckDB SQL") + 1
    except ValueError:
        insert_at = ws.max_column + 1
    ws.insert_cols(insert_at, amount=len(NEW_BACKEND_NAMES))
    for i, bn in enumerate(NEW_BACKEND_NAMES):
        c = insert_at + i
        ws.cell(row=1, column=c).value = f"{bn} rows"
        for q in QUERIES:
            r = find_q_row(ws, q)
            if r is None:
                continue
            rec = new_recs[bn].get(q)
            if rec and rec["rows"] is not None:
                ws.cell(row=r, column=c).value = rec["rows"]

    note_row = ws.max_row + 1
    ws.cell(row=note_row, column=1).value = "Note"
    ws.cell(row=note_row, column=2).value = (
        "CB+GE / CR+GE: teacher's month-level EE Group Encoding (84 cardinality). "
        "Q1 cutoff rounded to 1998-08-31 (was 1998-09-02); Q3 cutoff rounded "
        "to 1995-04-01 (was 1995-03-15); Q14/Q15 use month boundaries directly; "
        "Q6 uses year-level GE. SQL ground-truth comparison uses same rounded "
        "cutoffs (rows/aggregates differ from base-CB column by design)."
    )

    wb.save(XLSX)
    print(f"[saved] {XLSX}")


if __name__ == "__main__":
    main()
