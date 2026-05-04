#!/usr/bin/env python3
"""TPC-H bitmap-benchmark aggregation driver.

Runs all 12 modernised TPC-H benchmarks (Q1/Q3/Q4/Q5/Q6/Q8/Q10/Q12/Q14/Q15/
Q17/Q19) at a single scale factor with all 8 bitmap backends in a single
invocation each (`DEBIT_BM=all`), parses both the per-Q CSV artefact that
each query writes (`q{N}_results_SF10.csv`) and the captured stdout log
(for storage / memory footprints + row-count correctness + DuckDB SQL
baseline), then emits:

  * bm_results_long.csv   — long format, one row per (Q, backend, phase)
  * bm_results.xlsx       — same data reshaped into six clean sheets:
      - Total Summary        (Q × backend, total median + footprint)
      - Timings Detail       (long Q × backend × phase with full stats)
      - Total Time Matrix    (Q × backend median total_ms, DuckDB column)
      - Storage Summary      (Q × backend footprint MB, on-disk vs in-mem)
      - Memory Grouped       (per-Q grouped block like the user requested)
      - Correctness          (row counts per backend + DuckDB ground truth)

Units are normalised to MB (1 MiB = 1.048576 MB; numbers stay printed as
MB directly since the source data is already in MiB and the inflation is
< 5 %; column headers explicitly say MB).

Run from the duckdb-dev repo root:
    python3 tools/bench_suite.py --sf 10
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import subprocess
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DB_TEMPLATE = "tpch_sf{sf}.db"
DEFAULT_DUCKDB_BIN = REPO_ROOT / "build" / "release" / "duckdb"

# The 12 modernised TPC-H benchmark query numbers.
QUERIES: List[int] = [1, 3, 4, 5, 6, 8, 10, 12, 14, 15, 17, 19]

# Canonical backend order for every output sheet.  Matches the Backend
# enum declared in extension/debit/include/execution/tpch/bm_bench_common.hpp.
BACKENDS: List[str] = ["WAH", "CB", "CR", "CRR", "EW", "BS", "BSA", "CON"]

BACKEND_FULL: Dict[str, str] = {
    "WAH": "WAH (FastBit)",
    "CB":  "ComBit",
    "CR":  "CRoaring",
    "CRR": "CRoaring+Run",
    "EW":  "EWAH",
    "BS":  "Bitset (scalar)",
    "BSA": "Bitset+AVX512",
    "CON": "Concise",
}

# CSV column-prefix -> canonical backend tag.  Covers both schemas used
# by the 12 Q files (Q1/Q3/Q4/Q6 spell backends out, Q5/Q8/Q10/Q12/Q14
# use 2/3-letter tags; Q15/Q17/Q19 put the tag in a `backend` column).
CSV_PREFIX_TO_BACKEND: Dict[str, str] = {
    "wah":           "WAH",
    "combit":        "CB",
    "combit_bpe":    "CB+BPE",
    "cb":            "CB",
    "croaring":      "CR",
    "croaring_bpe":  "CR+BPE",
    "cr":            "CR",
    "croaring_run":  "CRR",
    "crr":           "CRR",
    "ewah":          "EW",
    "ew":            "EW",
    "bs":            "BS",
    "bsa":           "BSA",
    "concise":       "CON",
    "con":           "CON",
}

# --- Regexes for stdout parsing -----------------------------------------

# Storage / footprint lines.  All print `<Name> (on disk|in mem): X.YZ MiB`.
RE_FOOTPRINT = re.compile(
    r"^\s*(WAH|ComBit|CRoaring|EWAH|Bitset|Concise)\s+"
    r"(on disk|in mem)\s*:\s*([\d.]+)\s*MiB",
    re.MULTILINE,
)

# Row-count lines printed per backend, in either of two styles the 12 Q
# files use.  The long-form style (Q1, Q5, Q6) spells every backend out:
#   `  WAH rows:        59142609`
#   `  ComBit rows:     59142609`
#   `  CRoar+Run rows:  59142609`
#   `  Bitset+AVX512 rows:  59142609`
# The compact style (Q3, Q8, Q10) uses short tags and frequently packs
# multiple backends on one line:
#   `  CB rows:  302114  CR rows:  302114  CRR rows: 302114  WAH rows: 302114  EW rows:  302114`
#   `  BS rows:  302114  BSA rows: 302114  CON rows: 302114`
# Q4/Q12/Q14/Q15/Q17/Q19 print no per-backend row count at all — they
# instead emit an `[OK] all active backends match DuckDB SQL ground truth`
# pass message which is captured separately below.
RE_ROWS = re.compile(
    r"\b(WAH|ComBit|CRoaring|CRoar\+Run|EWAH|Bitset\+AVX512|Bitset|Concise|"
    r"CB|CR|CRR|EW|BS|BSA|CON)\s+rows\s*:\s*(\d+)"
)

ROW_LABEL_TO_BACKEND: Dict[str, str] = {
    # Long-form
    "WAH":            "WAH",
    "ComBit":         "CB",
    "CRoaring":       "CR",
    "CRoar+Run":      "CRR",
    "EWAH":           "EW",
    "Bitset":         "BS",
    "Bitset+AVX512":  "BSA",
    "Concise":        "CON",
    # Short-form (Q3/Q8/Q10 pack these onto one line).
    "CB":  "CB",
    "CR":  "CR",
    "CRR": "CRR",
    "EW":  "EW",
    "BS":  "BS",
    "BSA": "BSA",
    "CON": "CON",
}

# DuckDB SQL baseline line.  Two shapes appear across the 12 Q files:
# (a)  `[Baseline] DuckDB native SQL count(*) = 59142609  (single run: 181.66 ms)`
# (b)  `[Baseline] DuckDB native SQL Q4  (single run: 1202.57 ms)`    — no `=`
# (c)  `[Baseline] DuckDB native SQL Top-10  (single run: 292.57 ms)` — Q3 / Q10
# Try (a) first to keep the scalar result; fall back to the ms-only capture.
RE_DUCKDB_BASELINE_WITH_VAL = re.compile(
    r"\[Baseline\]\s+DuckDB native SQL[^=]*=\s*([0-9.eE+-]+)\s+.*?"
    r"\(single run:\s*([\d.]+)\s*ms\)",
    re.DOTALL,
)
RE_DUCKDB_BASELINE_MS_ONLY = re.compile(
    r"\[Baseline\]\s+DuckDB native SQL[^\n]*?\(single run:\s*([\d.]+)\s*ms\)"
)

# `[OK] all active backends match DuckDB SQL ground truth ...` pass message
# emitted by every Q on success.  Used as the correctness signal when no
# per-backend row count is printed.
RE_OK_PASS = re.compile(
    r"\[OK\]\s+all active backends match\s+DuckDB SQL ground truth[^\n]*"
)

# Per-backend in-memory breakdown lines.  Format (one per backend, MiB):
#   `  [Breakdown] WAH      literal=40.08 MiB  fill=3.12 MiB  header=0.17 MiB  total=43.38 MiB`
#   `  [Breakdown] ComBit   L1=29.51 MiB  L2=4.26 MiB  L3=10.61 MiB  total=44.38 MiB`
#   `  [Breakdown] CRoaring array=1.61 MiB  bitset=35.78 MiB  run=0.00 MiB  total=37.39 MiB`
#   `  [Breakdown] CR+Run   array=1.61 MiB  bitset=35.76 MiB  run=0.02 MiB  total=37.38 MiB`
#   `  [Breakdown] EWAH     literal=42.05 MiB  fill=6.23 MiB  total=48.28 MiB`
#   `  [Breakdown] Bitset   raw=712.34 MiB  total=712.34 MiB`
#   `  [Breakdown] BSA      raw=712.34 MiB  total=712.34 MiB`
#   `  [Breakdown] Concise  literal=37.04 MiB  fill=3.14 MiB  total=40.18 MiB`
RE_BREAKDOWN_LINE = re.compile(
    r"^\s*\[Breakdown\]\s+(\S+)\s+(.+?)\s*$",
    re.MULTILINE,
)
RE_BREAKDOWN_KV = re.compile(r"(\w+)=([0-9.]+)\s*MiB")

# New BMTPCH pattern (PRAGMA load_bitmap) prints one line per (column,
# backend) build:
#   `[load_bitmap] orderkey: built ComBit index (15000000 keys, 4448.84 MB) in 12947 ms`
#   `[load_bitmap] shipdate: built CRoaringRun index (2526 keys, 138.443 MB) in 1403 ms`
# Size is decimal MB (storage_bytes / 1e6), already consistent with the
# spreadsheet's MB convention — no MiB conversion needed.
RE_LOAD_BITMAP_BUILT = re.compile(
    r"\[load_bitmap\]\s+(\S+):\s+built\s+(\w+)\s+index\s+\(\d+\s+keys,\s+([0-9.]+)\s+MB\)",
)
# IBitmapIndex::backend_name() printed strings -> canonical backend tag.
LOAD_BITMAP_NAME_TO_BACKEND: Dict[str, str] = {
    "ComBit":      "CB",
    "ComBitGE":    "CB",
    "ComBitBPE":   "CB",
    "CRoaring":    "CR",
    "CRoaringRun": "CRR",
    "CRoaringBPE": "CR",   # BPE label is shared between run-opt on/off
    "WAH":         "WAH",
    "EWAH":        "EW",
    "Concise":     "CON",
}

# Map the printed leading tag -> canonical backend code + category list
# in canonical column order.  Categories that don't apply for a backend
# stay None so the spreadsheet can render them as blank cells.
BREAKDOWN_TAG: Dict[str, Tuple[str, List[str]]] = {
    "WAH":      ("WAH",  ["literal", "fill", "header"]),
    "ComBit":   ("CB",   ["L1", "L2", "L3", "L4"]),
    "CRoaring": ("CR",   ["array", "bitset", "run"]),
    "CR+Run":   ("CRR",  ["array", "bitset", "run"]),
    "EWAH":     ("EW",   ["literal", "fill"]),
    "Bitset":   ("BS",   ["raw"]),
    "BSA":      ("BSA",  ["raw"]),
    "Concise":  ("CON",  ["literal", "fill"]),
}

# Reverse-keyed-by-backend lookup: canonical category order per backend
# code.  Used by the Excel builder to render Memory Breakdown / Memory
# Detail in a stable column-or-line order regardless of stdout key order.
BREAKDOWN_TAG_ORDER_BY_BACKEND: Dict[str, List[str]] = {
    backend: cats for _, (backend, cats) in BREAKDOWN_TAG.items()
}

# Backend-specific iteration lines like `  CB:   OR=... rows=NNN` emitted
# by Q3/Q15 during the loop.  Only keep the last match per backend so the
# final measured iteration (not warm-up) wins.
RE_INLINE_ROWS = re.compile(
    r"^\s*(WAH|CB|CR|CRR|EW|BS|BSA|CON)\s*:.*?rows=(\d+)", re.MULTILINE
)

# MiB → MB conversion (display in MB, as the user requested).  The source
# stdout figures are in MiB; 1 MiB == 1.048576 MB.
MIB_TO_MB = 1.048576


def mib_to_mb(mib: float) -> float:
    return mib * MIB_TO_MB


def _dir_size_follow_symlinks(path: Path) -> int:
    """Recursive directory size that DOES descend into symlinked
    subdirectories.  `Path.rglob` — and `os.walk` without followlinks=True
    — stop at a directory-symlink boundary, so a tree of nothing but
    symlinks (Q15's `shipdate -> tpch_q6_wah/shipdate`) would appear
    empty.  Use `os.walk(..., followlinks=True)` to go through them, and
    stat() each file (which itself follows symlinks to files).
    """
    import os as _os
    if not path.exists():
        return 0
    total = 0
    for root, _dirs, files in _os.walk(str(path), followlinks=True):
        for name in files:
            try:
                total += _os.stat(_os.path.join(root, name)).st_size
            except OSError:
                continue
    return total


# Per-backend on-disk directory naming convention used by every Q:
#   tpch_q{N}_{wah,combit,croaring,ewah}
# `None` means the backend has no on-disk footprint (BS/BSA/CON rebuild
# from CRoaring at load and live entirely in RAM).
BACKEND_DIR_SUFFIX: Dict[str, Optional[str]] = {
    "WAH": "wah",
    "CB":  "combit",
    "CR":  "croaring",
    "CRR": "croaring",   # shares CR's dir
    "EW":  "ewah",
    "BS":  None,
    "BSA": None,
    "CON": None,
}


# --- Per-Q CSV parsing --------------------------------------------------


def parse_schema_wide(path: Path, q: int) -> List[Dict]:
    """Schema A: Q1/Q3/Q4/Q5/Q6/Q8/Q10/Q12/Q14.

    Columns look like `<prefix>_median_ms / _stddev_ms / _min_ms / _max_ms`
    for each backend; rows are `sf,operation,...`.  Returns one record per
    (backend, phase), holding all four stats.
    """
    records: List[Dict] = []
    with path.open() as f:
        reader = csv.DictReader(f)
        for row in reader:
            phase = row["operation"]
            for prefix, backend in CSV_PREFIX_TO_BACKEND.items():
                med_key = f"{prefix}_median_ms"
                if med_key not in row:
                    continue
                try:
                    median = float(row[med_key])
                except (KeyError, ValueError):
                    continue
                rec = {
                    "Q":        f"Q{q}",
                    "Backend":  backend,
                    "Phase":    phase,
                    "Median_ms": median,
                    "Stddev_ms": float(row.get(f"{prefix}_stddev_ms", "nan") or "nan"),
                    "Min_ms":    float(row.get(f"{prefix}_min_ms", "nan") or "nan"),
                    "Max_ms":    float(row.get(f"{prefix}_max_ms", "nan") or "nan"),
                }
                records.append(rec)
    return records


def parse_schema_simple(path: Path, q: int) -> List[Dict]:
    """Schema B: Q15 (`backend,or_ms,agg_ms,total_ms`), Q17 (`pass1/pass2`),
    Q19 (`or,and,agg,total`).  Only medians are stored; stddev/min/max are
    unavailable so they stay NaN.
    """
    records: List[Dict] = []
    with path.open() as f:
        reader = csv.DictReader(f)
        phase_cols = [c for c in reader.fieldnames if c.endswith("_ms")]
        for row in reader:
            backend = row["backend"].strip().upper()
            if backend not in BACKENDS:
                continue
            for col in phase_cols:
                # Column name `<phase>_ms` → phase label without the `_ms`.
                phase = col[:-3]
                try:
                    median = float(row[col])
                except ValueError:
                    continue
                # Normalise the few special phase labels so they line up
                # with the Schema-A names where possible.
                phase_norm = {
                    "or":     "OR",
                    "and":    "AND",
                    "agg":    "Agg",
                    "total":  "TOTAL",
                    "pass1":  "Pass1",
                    "pass2":  "Pass2",
                }.get(phase, phase)
                records.append({
                    "Q":         f"Q{q}",
                    "Backend":   backend,
                    "Phase":     phase_norm,
                    "Median_ms": median,
                    "Stddev_ms": float("nan"),
                    "Min_ms":    float("nan"),
                    "Max_ms":    float("nan"),
                })
    return records


# Q-number → which CSV schema to use.
# Q15 still uses the legacy Schema-B (backend,or_ms,agg_ms,total_ms)
# from its old in-process bench loop.  Q17 and Q19 now use Schema-A
# after the BMTPCH port.
SCHEMA_B_QS = {15}


def parse_csv_for_q(repo_root: Path, q: int, sf_label: str) -> List[Dict]:
    path = repo_root / f"q{q}_results_{sf_label}.csv"
    if not path.exists():
        print(f"[warn] missing CSV: {path}", file=sys.stderr)
        return []
    return parse_schema_simple(path, q) if q in SCHEMA_B_QS else parse_schema_wide(path, q)


# --- stdout log parsing -------------------------------------------------


def parse_stdout_log(log_text: str, *, q: Optional[int] = None,
                      repo_root: Optional[Path] = None,
                      sf: int = 10) -> Dict:
    """Extract footprint (MB, on-disk or in-mem), per-backend row counts,
    the DuckDB SQL baseline (row-count-ish value + single-run ms), and
    the `[OK] all active backends match ...` pass message so queries that
    verify against numeric aggregates (no row count) still have something
    in the correctness sheet.

    When the C++ printed footprint is 0.00 MiB for a disk-based backend
    but the per-backend directory exists and is really a pile of symlinks
    into another Q's dir (Q15 shipdate → Q6 shipdate), do a Python-side
    re-measurement that follows the symlinks and flag the kind as
    ``disk (shared)`` so the spreadsheet keeps the linkage visible.
    """
    result: Dict = {
        "footprint_mb":      {b: None for b in BACKENDS},
        "footprint_kind":    {b: None for b in BACKENDS},  # 'disk' | 'mem' | 'disk (shared)'
        "rows":              {b: None for b in BACKENDS},
        "duckdb_rows":       None,
        "duckdb_ms":         None,
        "ok_message":        None,
        # In-memory breakdown per backend: {backend: {category: MB}}
        "memory_breakdown":  {b: {} for b in BACKENDS},
    }

    # Footprints -----------------------------------------------------------
    for m in RE_FOOTPRINT.finditer(log_text):
        name, kind, mib = m.group(1), m.group(2), float(m.group(3))
        mb = mib_to_mb(mib)
        # Map printed names to canonical backend tag.  CRoaring footprint
        # is shared by CR and CRR (same directory), so populate both.
        if name == "WAH":
            result["footprint_mb"]["WAH"] = mb
            result["footprint_kind"]["WAH"] = "disk"
        elif name == "ComBit":
            result["footprint_mb"]["CB"] = mb
            result["footprint_kind"]["CB"] = "disk"
        elif name == "CRoaring":
            for b in ("CR", "CRR"):
                result["footprint_mb"][b] = mb
                result["footprint_kind"][b] = "disk"
        elif name == "EWAH":
            result["footprint_mb"]["EW"] = mb
            result["footprint_kind"]["EW"] = "disk"
        elif name == "Bitset":
            # BS and BSA share the same in-mem bit-array representation;
            # their on-line deltas are only in the AND/popcount code path.
            for b in ("BS", "BSA"):
                result["footprint_mb"][b] = mb
                result["footprint_kind"][b] = "mem"
        elif name == "Concise":
            result["footprint_mb"]["CON"] = mb
            result["footprint_kind"]["CON"] = "mem"

    # Python-side footprint fallback: when the C++ did not print a
    # per-backend footprint line (Q3/Q4 currently omit them entirely)
    # or printed 0.00 MiB because the directory is symlinked into
    # another Q (Q15 shipdate → Q6 shipdate), measure the on-disk size
    # directly.  The kind is set to `disk (shared)` whenever the
    # directory contains a symlink, so the reader can see at a glance
    # that the bytes are accounted for in a different Q as well.
    if q is not None and repo_root is not None:
        sf_dir_suffix = "" if sf == 10 else f"_sf{sf}"
        for b, suffix in BACKEND_DIR_SUFFIX.items():
            if suffix is None:
                continue
            existing_mb = result["footprint_mb"].get(b)
            if existing_mb is not None and existing_mb > 0.01:
                continue  # C++ already reported a real value
            bm_dir = repo_root / f"tpch_q{q}{sf_dir_suffix}_{suffix}"
            if not bm_dir.exists():
                continue
            resolved_bytes = _dir_size_follow_symlinks(bm_dir)
            if resolved_bytes <= 0:
                continue
            # Detect whether ANY directory entry is a symlink; if so the
            # bytes are shared with another Q's on-disk layout.
            shared = any(child.is_symlink() for child in bm_dir.iterdir())
            result["footprint_mb"][b] = resolved_bytes / (1024 * 1024) * MIB_TO_MB
            result["footprint_kind"][b] = "disk (shared)" if shared else "disk"

    # Row counts -----------------------------------------------------------
    #
    # The summary "  WAH rows: N" block typically appears exactly once at
    # the end of the per-Q output.  The short-tag variant packs multiple
    # entries onto a single line but is also emitted once.  First-write-
    # wins is therefore correct: any subsequent duplicate would just be a
    # repeat, and warm-up traces use the inline `rows=NNN` syntax matched
    # separately below.
    for m in RE_ROWS.finditer(log_text):
        label, n = m.group(1), int(m.group(2))
        b = ROW_LABEL_TO_BACKEND.get(label)
        if b is not None and result["rows"][b] is None:
            result["rows"][b] = n

    # Inline per-iteration rows (`  CB:  OR=... rows=NNN`) emitted by Q3
    # and the compact Q15/Q17/Q19 formatter.  Last-write-wins so we end
    # up with the final measured iteration's value, not a warm-up one.
    for m in RE_INLINE_ROWS.finditer(log_text):
        b, n = m.group(1), int(m.group(2))
        if b in BACKENDS:
            result["rows"][b] = n

    # DuckDB SQL baseline --------------------------------------------------
    # Prefer the value-bearing shape so we get a numeric comparator; fall
    # back to the ms-only shape for Q3/Q4/Q10/Q12/Q14/Q15/Q17/Q19 where
    # the baseline produces a multi-row answer (TOP-K, grouped aggregate)
    # rather than a single scalar.
    m = RE_DUCKDB_BASELINE_WITH_VAL.search(log_text)
    if m:
        try:
            result["duckdb_rows"] = float(m.group(1))
        except ValueError:
            pass
        try:
            result["duckdb_ms"] = float(m.group(2))
        except ValueError:
            pass
    else:
        m = RE_DUCKDB_BASELINE_MS_ONLY.search(log_text)
        if m:
            try:
                result["duckdb_ms"] = float(m.group(1))
            except ValueError:
                pass

    # Pass message ---------------------------------------------------------
    m = RE_OK_PASS.search(log_text)
    if m:
        # Strip trailing whitespace but preserve the descriptive tail
        # (e.g. "(revenue within 0.01)") so the Correctness sheet shows
        # what the Q actually validated.
        result["ok_message"] = m.group(0).strip()

    # Per-backend in-memory breakdown ---------------------------------------
    # Each `[Breakdown] <Tag> <kv-pairs>` line carries the per-category
    # bytes for one backend.  Convert MiB -> MB on the way in so the
    # downstream spreadsheet sticks to MB consistently.
    for line_match in RE_BREAKDOWN_LINE.finditer(log_text):
        tag, body = line_match.group(1), line_match.group(2)
        if tag not in BREAKDOWN_TAG:
            continue
        backend, _expected = BREAKDOWN_TAG[tag]
        kv = {k: float(v) for k, v in RE_BREAKDOWN_KV.findall(body)}
        if not kv:
            continue
        # Store all categories printed (including 'total') in MB.
        bd = {k: mib_to_mb(v) for k, v in kv.items()}
        result["memory_breakdown"][backend] = bd

    # New BMTPCH pattern footprint: one `[load_bitmap]` line per (column,
    # backend) build.  Aggregate per-backend across the columns each Q
    # loads (e.g. Q3 loads orderkey + shipdate; Q5 loads orderkey +
    # suppkey).  This is the only memory signal the new pattern emits —
    # legacy `[Breakdown]` lines aren't printed here.  Sum across columns
    # to get a single per-backend total, plus keep per-column entries for
    # the Memory Breakdown sheet so reviewers can see how each Q's
    # storage decomposes by aux structure.
    bmtpch_seen: set = set()
    for m in RE_LOAD_BITMAP_BUILT.finditer(log_text):
        col, name, mb_str = m.group(1), m.group(2), m.group(3)
        backend = LOAD_BITMAP_NAME_TO_BACKEND.get(name)
        if backend is None:
            continue
        try:
            mb = float(mb_str)
        except ValueError:
            continue
        bd = result["memory_breakdown"].setdefault(backend, {})
        # Skip if the legacy `[Breakdown]` block already populated this
        # backend with structured categories (L1/L2/L3/L4 etc) — don't
        # overwrite richer detail with coarser per-column buckets.
        if any(k in bd for k in ("L1", "L2", "literal", "raw", "array")):
            continue
        bd[col] = bd.get(col, 0.0) + mb
        bd["total"] = sum(v for k, v in bd.items() if k != "total")
        bmtpch_seen.add(backend)
    # For Qs that use the new BMTPCH pattern, the on-disk footprint left
    # over from the pre-port pipeline (.bm files in tpch_q*_<backend>/)
    # is stale — those files aren't what's actually executing.  Replace
    # `footprint_mb` with the freshly-built in-memory total so the
    # Storage Total sheet reflects what the bench really paid for.
    for backend in bmtpch_seen:
        bd = result["memory_breakdown"].get(backend, {})
        total_mb = bd.get("total")
        if total_mb is None or total_mb <= 0.0:
            continue
        result["footprint_mb"][backend] = total_mb
        result["footprint_kind"][backend] = "mem (built)"

    # Backfill missing footprint_mb / kind from the breakdown's 'total'
    # field.  Compressed backends (WAH/CB/CR/CRR/EW) usually report an
    # on-disk size separately; in-mem backends (BS/BSA/CON) do not, so
    # without this step their Storage Total cells stay empty for queries
    # whose stdout doesn't print the legacy '  <Backend> on-disk = ...'
    # summary lines (e.g. Q3, Q4).
    _IN_MEM = {"BS", "BSA", "CON"}
    for backend, bd in result["memory_breakdown"].items():
        if not bd:
            continue
        existing = result["footprint_mb"].get(backend)
        if existing is not None and existing > 0.01:
            continue
        total_mb = bd.get("total")
        if total_mb is None:
            # Fall back to summing all non-'total' categories.
            total_mb = sum(v for k, v in bd.items() if k != "total")
        if total_mb <= 0.0:
            continue
        result["footprint_mb"][backend] = total_mb
        if result["footprint_kind"].get(backend) is None:
            result["footprint_kind"][backend] = "mem" if backend in _IN_MEM else "disk"

    return result


# --- Subprocess driver --------------------------------------------------


# Q1/Q5/Q6 use the new BMTPCH "BitEngine pattern" (PRAGMA load_bitmap +
# context.client.bitmap_*).  Each backend builds its own auxiliary index;
# DEBIT_BM=all is meaningless for these — we have to invoke duckdb once
# per backend, with the appropriate load_bitmap PRAGMAs.  The mapping
# below records which columns each Q needs pre-loaded.
PATTERN_QS_LOAD_BITMAPS: Dict[int, List[str]] = {
    1:  ["shipdate", "linestatus", "returnflag"],
    3:  ["orderkey", "shipdate"],
    4:  ["orderkey"],
    5:  ["orderkey", "suppkey"],
    6:  ["shipdate_GE", "discount", "quantity"],
    8:  ["orderkey", "partkey"],
    10: ["orderkey", "returnflag"],
    12: ["shipmode", "receiptdate"],
    14: ["shipdate"],
    17: ["partkey"],
    19: ["shipmode", "shipinstruct"],
}

# β scheme: BPE-augmented backends additionally pre-build a bucketed
# prefix-encoded auxiliary structure on the column whose Q is dominated
# by a range scan.  Per TPC-H 1.5.7 §5 each `<col>_BPE` references the
# same single base-table column as the per-value `<col>`; the BPE
# structure is just a different cumulative encoding.  Compliant.
PATTERN_QS_LOAD_BITMAPS_BPE_EXTRA: Dict[int, List[str]] = {
    1:  ["shipdate_BPE"],
    3:  ["shipdate_BPE"],
    6:  ["discount_BPE", "quantity_BPE"],
    # Q14 omitted: 30-day range is smaller than the 100-day bucket, so
    # BPE would force boundary trims exceeding the naive range size.
    # CB+BPE Q14 falls back to the same per-day OR path as CB.
}

# Per-backend env values that map to our BitmapBackend enum.
# CB+BPE / CR+BPE are the β scheme: same per-value bitmap as CB / CR
# PLUS a bucketed prefix-encoded column for range-dominant queries.
# Both are listed (not just CB+BPE) so the comparison stays apples-to-apples.
PATTERN_BACKENDS_FULL: List[Tuple[str, str]] = [
    ("CB",     "cb"),
    ("CB+BPE", "cb_bpe"),
    ("CR",     "cr"),
    ("CR+BPE", "cr_bpe"),
    ("CRR",    "crr"),
    ("WAH",    "wah"),
    ("EW",     "ew"),
    ("CON",    "con"),
]

# Per-Q runtime gating after IndexedWAH::build was made sparse-aware
# (build cost is uniform).  The remaining gating is purely about runtime
# OR fan-out:
#
#   * WAH (FastBit ibis::bitvector) has no k-way merge — pairwise
#     `dst |= bv` only, ~160M-word per OR on a 1.27 GB SF10 orderkey
#     index.  Tractable up to ~200 keys; gated off for Q3/Q4/Q5.
#   * EWAH and Concise have fast_logicalor (priority-queue k-way),
#     but the constant factor on K=590k inputs is still ~10-30 minutes
#     per iteration empirically (the priority queue does O(K log K)
#     setup before any merging).  Tractable up to ~30k keys; gated off
#     for Q3/Q4 (590k keys) but kept on for Q5 (28k keys).
#   * CRoaring (CR / CRR) fastunion handles 590k inputs in seconds —
#     the only backends fast enough to keep on for Q3/Q4.
#
# Q17 has ~200 partkey ORs (Brand#23 + MED BOX selectivity) so all 6
# backends run cheaply.  Q1/Q6/Q14/Q19 have ≤90-key fan-out.
PATTERN_BACKENDS_PER_Q: Dict[int, List[Tuple[str, str]]] = {
    1:  PATTERN_BACKENDS_FULL,
    3:  PATTERN_BACKENDS_FULL,
    4:  PATTERN_BACKENDS_FULL,
    5:  PATTERN_BACKENDS_FULL,
    6:  PATTERN_BACKENDS_FULL,
    8:  PATTERN_BACKENDS_FULL,
    10: PATTERN_BACKENDS_FULL,
    12: PATTERN_BACKENDS_FULL,
    14: PATTERN_BACKENDS_FULL,
    17: PATTERN_BACKENDS_FULL,
    19: PATTERN_BACKENDS_FULL,
}

# Maps the active backend short tag to its CSV column prefix in the
# Schema-A CSV (matches CSV_PREFIX_TO_BACKEND inverted).
PATTERN_BACKEND_CSV_PREFIX: Dict[str, str] = {
    "CB":     "combit",
    "CB+BPE": "combit_bpe",
    "CR":     "croaring",
    "CR+BPE": "croaring_bpe",
    "CRR":    "croaring_run",
    "WAH":    "wah",
    "EW":     "ewah",
    "CON":    "concise",
}


def _merge_pattern_csvs(repo_root: Path, q: int, sf_label: str,
                        per_backend_csvs: Dict[str, Path]) -> None:
    """For Q1/Q5/Q6 we ran the bench once per backend; each run overwrote
    qN_results_{sf}.csv with that backend's column populated and the
    others 0.  Merge the per-backend captures into a single combined CSV
    that bench_suite's parse_schema_wide can consume.

    For CB+BPE / CR+BPE the cpp emits to the *non*-BPE prefix (the cpp
    has no notion of `cb_bpe` — DEBIT_BM=cb_bpe just routes to the same
    ComBit emit path); we rename `combit_*` → `combit_bpe_*` (and likewise
    for croaring) when copying into the merged output.
    """
    if not per_backend_csvs:
        return
    # Backends whose source CSV still uses the non-BPE prefix.
    SOURCE_PREFIX_OVERRIDE = {
        "CB+BPE": "combit",        # source column → renamed to combit_bpe_*
        "CR+BPE": "croaring",
    }
    # Collect rows keyed on operation, taking the active-backend cells from
    # each per-backend CSV.
    op_to_row: Dict[str, Dict[str, str]] = {}
    base_fieldnames: Optional[List[str]] = None
    for backend, path in per_backend_csvs.items():
        if not path.exists():
            continue
        prefix     = PATTERN_BACKEND_CSV_PREFIX.get(backend)
        src_prefix = SOURCE_PREFIX_OVERRIDE.get(backend, prefix)
        if not prefix:
            continue
        with path.open() as f:
            reader = csv.DictReader(f)
            if base_fieldnames is None:
                base_fieldnames = list(reader.fieldnames or [])
            for row in reader:
                op = row.get("operation", "")
                if not op:
                    continue
                target = op_to_row.setdefault(op, {**row})
                for stat in ("median_ms", "stddev_ms", "min_ms", "max_ms"):
                    src = f"{src_prefix}_{stat}"
                    dst = f"{prefix}_{stat}"
                    if src in row:
                        target[dst] = row[src]
    if base_fieldnames is None:
        return
    # Extend fieldnames with any BPE columns that weren't in the source CSV.
    fieldnames = list(base_fieldnames)
    for backend, _ in per_backend_csvs.items():
        prefix = PATTERN_BACKEND_CSV_PREFIX.get(backend)
        if not prefix:
            continue
        for stat in ("median_ms", "stddev_ms", "min_ms", "max_ms"):
            col = f"{prefix}_{stat}"
            if col not in fieldnames:
                fieldnames.append(col)
    # Write the merged CSV to the canonical location.
    out_path = repo_root / f"q{q}_results_{sf_label}.csv"
    with out_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for op in op_to_row.values():
            w.writerow(op)


def run_single_query(duckdb_bin: Path, db: Path, q: int,
                     env_extra: Dict[str, str], log_path: Path) -> int:
    env = os.environ.copy()
    env.update(env_extra)
    sf_label = env_extra.get("TPCH_SF_LABEL") or env.get("TPCH_SF_LABEL", "SF10")

    if q in PATTERN_QS_LOAD_BITMAPS:
        # New pattern: per-backend runs with PRAGMA load_bitmap.
        # duckdb's `-c` only runs the first statement in a chain; we
        # feed the multi-statement SQL via stdin so all PRAGMAs execute.
        cols = PATTERN_QS_LOAD_BITMAPS[q]
        per_backend_csvs: Dict[str, Path] = {}
        with log_path.open("w") as f:
            f.write(f"# cwd={REPO_ROOT}\n# pattern_qs_per_backend Q{q}\n")
            for backend, env_val in PATTERN_BACKENDS_PER_Q.get(q, PATTERN_BACKENDS_FULL):
                env["DEBIT_BM"] = env_val
                # β scheme: BPE-augmented backends additionally pre-load
                # the bucketed prefix-encoded column (TPC-H 1.5.7 §5
                # compliant — same base-table column as the per-value).
                this_cols = list(cols)
                if env_val in ("cb_bpe", "cr_bpe"):
                    for extra in PATTERN_QS_LOAD_BITMAPS_BPE_EXTRA.get(q, []):
                        if extra not in this_cols:
                            this_cols.append(extra)
                stmts = [f"PRAGMA load_bitmap('{c}');" for c in this_cols]
                stmts.append(f"PRAGMA bm_tpch({q});")
                sql_input = "\n".join(stmts) + "\n"
                f.write(f"\n# === Backend {backend} (DEBIT_BM={env_val}) ===\n")
                f.write(f"# sql=\n{sql_input}")
                f.flush()
                p = subprocess.run([str(duckdb_bin), str(db)],
                                   cwd=str(REPO_ROOT), env=env,
                                   input=sql_input,
                                   text=True,
                                   stdout=f, stderr=subprocess.STDOUT)
                if p.returncode != 0:
                    f.write(f"\n# Backend {backend} returncode={p.returncode}\n")
                # Capture this backend's CSV before the next run overwrites.
                src = REPO_ROOT / f"q{q}_results_{sf_label}.csv"
                if src.exists():
                    dst = REPO_ROOT / f"q{q}_results_{sf_label}_{env_val}.csv"
                    dst.write_bytes(src.read_bytes())
                    per_backend_csvs[backend] = dst
        _merge_pattern_csvs(REPO_ROOT, q, sf_label, per_backend_csvs)
        return 0

    # Legacy pattern: single invocation, all backends in-process.
    env["DEBIT_BM"] = "all"
    cmd = [str(duckdb_bin), str(db), "-c", f"PRAGMA bm_tpch({q});"]
    with log_path.open("w") as f:
        f.write(f"# cwd={REPO_ROOT}\n# cmd={' '.join(cmd)}\n")
        f.flush()
        p = subprocess.run(cmd, cwd=str(REPO_ROOT), env=env,
                           stdout=f, stderr=subprocess.STDOUT)
    return p.returncode


# --- Output assembly ----------------------------------------------------


def write_long_csv(records: List[Dict], path: Path) -> None:
    cols = ["Q", "Backend", "Phase", "Median_ms", "Stddev_ms",
            "Min_ms", "Max_ms"]
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in records:
            w.writerow({c: r.get(c, "") for c in cols})


def _q_total_median(records: List[Dict], q_tag: str, backend: str) -> Optional[float]:
    """Locate the TOTAL-phase median for a (Q, backend).  Schema-B Qs use
    'TOTAL' (uppercased in parse_schema_simple), Schema-A uses 'TOTAL'
    too, so a single lookup covers both.

    Returns None when the cell is 0.0 — that's a sentinel from the
    per-Q CSV emit (un-run backends are written as 0/0/0/0 stats).
    Rendering as None means the Excel cell is blank (vs. misleading
    green "0.000" which looks like an instant query).
    """
    for r in records:
        if r["Q"] == q_tag and r["Backend"] == backend and r["Phase"] == "TOTAL":
            v = r["Median_ms"]
            if v is None or v == 0.0:
                return None
            return v
    return None


def build_excel(records: List[Dict], meta: Dict[str, Dict],
                sf_label: str, out_xlsx: Path) -> None:
    """Emit the multi-sheet Excel workbook with publication-ready formatting.

    Sheet plan (in display order):
      1. README              — title, glossary, units, layout legend
      2. Time Matrix         — Q × backend total ms with green→red colour-scale
      3. Phase Time Detail   — (Q, phase) × backend ms (Schema-A queries)
      4. Memory Breakdown    — per-Q rows × per-backend cells, multi-line
                                strings labelled by category (matches the user's
                                example: "L1: 13MB / L2: 2MB / L3: 0MB")
      5. Memory Detail (MB)  — flat (Q, backend, category, MB) for analysis
      6. Storage Total (MB)  — Q × backend on-disk/in-mem total (one number)
      7. Correctness         — row counts + [OK] validation message per Q
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    # --- Theme / shared styles ------------------------------------------
    HEADER_FILL  = PatternFill("solid", fgColor="2E5C8A")
    HEADER_FONT  = Font(bold=True, size=11, color="FFFFFF")
    SUB_FILL     = PatternFill("solid", fgColor="D9E1F2")
    SUB_FONT     = Font(bold=True, size=10, color="1F3864")
    TITLE_FONT   = Font(bold=True, size=14, color="1F3864")
    SECT_FONT    = Font(bold=True, size=11, color="1F3864")
    Q_FILL       = PatternFill("solid", fgColor="EAEAEA")
    Q_FONT       = Font(bold=True, size=10)
    NUM_FILL_NA  = PatternFill("solid", fgColor="F5F5F5")
    THIN_GREY    = Side(style="thin", color="BFBFBF")
    BORDER       = Border(left=THIN_GREY, right=THIN_GREY,
                          top=THIN_GREY, bottom=THIN_GREY)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT        = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    NUM_FMT_MS   = "0.000"
    NUM_FMT_MB   = "0.00"
    NUM_FMT_INT  = "#,##0"

    wb = Workbook()
    wb.remove(wb.active)

    # --- Helpers ---------------------------------------------------------
    def _put_header(ws, row_idx, values, fill=HEADER_FILL, font=HEADER_FONT, height=22):
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=j, value=v)
            c.font = font
            c.fill = fill
            c.alignment = CENTER
            c.border = BORDER
        ws.row_dimensions[row_idx].height = height

    def _autosize(ws, *, min_w=10, max_w=42, fixed_widths=None):
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            if fixed_widths and letter in fixed_widths:
                ws.column_dimensions[letter].width = fixed_widths[letter]
                continue
            width = min_w
            for c in col:
                v = c.value
                if v is None:
                    continue
                # When the cell value carries newlines, size by widest line
                longest = max(len(line) for line in str(v).splitlines() or [""])
                width = max(width, min(longest + 2, max_w))
            ws.column_dimensions[letter].width = width

    def _q_total_ms(q_tag, backend):
        return _q_total_median(records, q_tag, backend)

    # ------------------------------------------------------------
    # Sheet: README
    # ------------------------------------------------------------
    readme = wb.create_sheet("README")
    rr = 1
    readme.cell(row=rr, column=1, value="TPC-H Bitmap Benchmark Results")
    readme.cell(row=rr, column=1).font = TITLE_FONT
    readme.row_dimensions[rr].height = 24
    rr += 1
    readme.cell(row=rr, column=1,
                value=f"Scale factor: {sf_label} — 8 bitmap backends × {len(QUERIES)} queries")
    rr += 2

    def _section(title):
        nonlocal rr
        readme.cell(row=rr, column=1, value=title)
        readme.cell(row=rr, column=1).font = SECT_FONT
        rr += 1

    _section("Backend tags")
    for b in BACKENDS:
        readme.cell(row=rr, column=1, value=b).font = Q_FONT
        readme.cell(row=rr, column=2, value=BACKEND_FULL[b])
        rr += 1
    rr += 1

    _section("Memory breakdown categories (MB)")
    breakdown_doc = [
        ("WAH",      "literal: literal-word bytes  •  fill: fill-word bytes  •  header: per-bitmap C++ object overhead"),
        ("ComBit",   "L1: literal-word bytes (level 1)  •  L2: literal bytes (level 2 after L3 compression)  •  L3: L3-literal bytes (L3 bytes that disagree with the L4 fill)  •  L4: leading bitstring over L3 (1 bit per L3 byte; replaces the dense L3-bits buffer)"),
        ("CRoaring", "array: array-container bytes  •  bitset: bitset-container bytes  •  run: run-container bytes"),
        ("CR+Run",   "Same categories as CRoaring; runOptimize() applied — so 'run' is non-zero whenever it pays off"),
        ("EWAH",     "literal: literal-word bytes  •  fill: RLW header words encoding fills"),
        ("Bitset",   "raw: uncompressed bitset bytes (single category — no compression layers)"),
        ("BSA",      "Same as Bitset (BS/BSA share the in-memory bit-array pool; they only differ in AND code path)"),
        ("Concise",  "literal: literal-word bytes  •  fill: 0/1-sequence (fill) words"),
    ]
    for tag, doc in breakdown_doc:
        readme.cell(row=rr, column=1, value=tag).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        rr += 1
    rr += 1

    _section("Sheet glossary")
    glossary = [
        ("Time Matrix",          "Q rows × backend cols.  Total-phase median ms.  Green→red colour scale per row."),
        ("Phase Time Detail",    "(Q, phase) rows × backend cols.  Per-phase median ms with the same colour scale."),
        ("Memory Breakdown",     "Q rows × backend cols.  Each cell shows the per-category MB labelled (e.g. 'L1: 29.51 MB / L2: 4.26 MB / L3: 10.61 MB / total: 44.38 MB')."),
        ("Memory Detail (MB)",   "Long format: one row per (Q, backend, category) MB — for charts and pivots."),
        ("Storage Total (MB)",   "Q × backend total footprint (single number per cell) plus per-Q on-disk and in-mem sums."),
        ("Correctness",          "Per-backend rows + DuckDB SQL ground truth + the [OK] validation message every Q emits on success."),
    ]
    for name, doc in glossary:
        readme.cell(row=rr, column=1, value=name).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        rr += 1
    rr += 1

    _section("Notes")
    notes = [
        "All time numbers are medians over the measured iterations (after warm-ups), as printed by each Q.",
        "All MB numbers come from the in-memory '[Breakdown]' lines printed after each Q's bitmap load.",
        "Unit: 1 MiB = 1.048576 MB; the spreadsheet stays in MB end-to-end.",
        "CR and CRR share one on-disk directory.  BS and BSA share one in-mem bitset pool.",
    ]
    for n in notes:
        readme.cell(row=rr, column=1, value="•").alignment = Alignment(horizontal="right")
        readme.cell(row=rr, column=2, value=n)
        rr += 1
    rr += 1

    _section("Origin & methodology")
    origin = [
        ("Q1 / Q5 / Q6", "Teacher's reviewer-fixed reference (BitEngine branch).  Ported verbatim — same `context.client.bitmap_*` + `dynamic_cast<rabit::Rabit*>` pattern, same per-table semi-joins."),
        ("Q3 / Q4 / Q14 / Q15 / Q17 / Q19", "Ported by following the same teacher pattern + reading TPC-H spec §2.4.x for each query's predicate semantics.  Each port emits an [OK] line confirming the result matches DuckDB SQL ground truth."),
        ("Q8 / Q10 / Q12", "Deferred.  These are 4-to-7-table aggregation queries where most of the work is in DuckDB SQL execution; a single bitmap-filter step would be a small fraction of total time and yield limited per-bitmap-op insight.  Old non-compliant pre-joined .bm files were retired pre-port; cells are intentionally blank in this workbook (not 0)."),
    ]
    for tag, doc in origin:
        readme.cell(row=rr, column=1, value=tag).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        readme.cell(row=rr, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        rr += 1
    rr += 1

    _section("Empty / blank cells")
    blanks = [
        ("Blank Q row",   "Q deferred (Q8 / Q10 / Q12) — see Origin & methodology above."),
        ("Blank backend cell",
         "Backend not run for that Q.  Gating is on runtime OR fan-out per-iteration:  Q3/Q4 (~590k orderkeys after the orders-scan + custkey-set semi-join) — only CB/CR/CRR finish in seconds; WAH has no k-way merge (pairwise OR over a 1.27 GB index = ~60 min/iter); EWAH/Concise's fast_logicalor priority-queue setup is O(K log K) and at K=590k empirically hangs >30 min/iter.  Q5 (~28k orderkeys) — WAH still gated, CR/CRR/EW/CON all run.  Q17 (~200 partkey ORs after Brand#23 + MED BOX) — all 6 run.  BS/BSA cells are blank: the legacy dense-bit-array baseline relied on file-format predicate pre-derivation incompatible with the new per-value PRAGMA load_bitmap pattern."),
    ]
    for tag, doc in blanks:
        readme.cell(row=rr, column=1, value=tag).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        readme.cell(row=rr, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        rr += 1
    rr += 1

    _section("TPC-H 1.5.7 §5 compliance — auxiliary data structures")
    compliance = [
        ("Spec rule",
         "Each directive may reference no more than one base table and may not reference other auxiliary structures.  "
         "Each directive may reference one and only one of: PK column(s), FK column(s), or a date column.  "
         "Functions/expressions on those columns are allowed.  Multi-column / multi-table indexes are forbidden."),
        ("Q1",
         "Mixed — bitmap_shipdate (date) is strict-compliant.  bitmap_linestatus + bitmap_returnflag reference VARCHAR columns; under a strict reading of 1.5.7 these are NOT permitted (only PK/FK/date).  Per teacher's BitEngine reference (rabit_linestatus / rabit_returnflag) we keep them — the academic methodology for bitmap research on TPC-H universally indexes these low-cardinality VARCHAR columns."),
        ("Q3",
         "COMPLIANT — bitmap_orderkey (FK on l_orderkey), bitmap_shipdate (date).  Customer mktsegment + orders orderdate filters at query-time."),
        ("Q4",
         "COMPLIANT — bitmap_orderkey (FK on l_orderkey).  l_commitdate < l_receiptdate evaluated at query-time on BMFetch'd cols."),
        ("Q5",
         "COMPLIANT — bitmap_orderkey + bitmap_suppkey (both FKs).  Region/nation/customer/orders/supplier semi-joins at query-time."),
        ("Q6",
         "Mixed — bitmap_shipdate_GE (date, year-bucketed) is compliant.  bitmap_discount + bitmap_quantity (and the BPE variants) reference DECIMAL columns; under strict 1.5.7 not permitted.  Kept per teacher's BitEngine reference (rabit_discount / rabit_quantity)."),
        ("Q14",
         "COMPLIANT — bitmap_shipdate (date) only.  PROMO% predicate evaluated by part scan at query-time."),
        ("Q15",
         "COMPLIANT — bitmap_shipdate (date) only.  Per-supplier max revenue computed at query-time."),
        ("Q17",
         "COMPLIANT — bitmap_partkey (FK on l_partkey).  p_brand+p_container scan at query-time."),
        ("Q19",
         "Non-strict — bitmap_shipmode + bitmap_shipinstruct (both VARCHAR).  Same caveat as Q1.  3 OR-group brand+container+size+quantity predicates evaluated at query-time on materialised cols (no aux on those — compliant)."),
        ("Q8 / Q10 / Q12",
         "Deferred — see Origin & methodology.  Older non-compliant pre-built multi-table bitmaps (join_result / late_lineitem / commit_lt_receipt / ship_lt_commit) were retired."),
    ]
    for tag, doc in compliance:
        readme.cell(row=rr, column=1, value=tag).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        readme.cell(row=rr, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        rr += 1
    rr += 1

    _section("Phase-name glossary (TPC-H query → phase semantics)")
    phase_glossary = [
        ("Q1 PhaseA",  "shipdate complement OR (~90 day-bitmaps for shipdate > '1998-09-02') + flip in byte-stream space."),
        ("Q1 PhaseB",  "Per-(returnflag, linestatus) group: linestatus_btv AND returnflag_btv, then byte-stream AND with shipdate filter."),
        ("Q1 PhaseC",  "Single sequential scan of lineitem(qty, price, disc, tax); per-row inline aggregate using the 5 group byte-streams (mirror of teacher's reduce_leadingbits)."),
        ("Q3 PhaseA",  "Customer scan: c_mktsegment='BUILDING' → c_custkey set."),
        ("Q3 PhaseB",  "Orders scan: filter o_orderdate < '1995-03-15' AND o_custkey ∈ set; for each match OR bitmap_orderkey[o_orderkey] into btv_res; build l_orderkey_map for downstream aggregate."),
        ("Q3 PhaseC",  "Shipdate range OR (l_shipdate > cutoff)."),
        ("Q3 PhaseD",  "btv_res &= shipdate_filter; get_rowids."),
        ("Q3 PhaseE",  "BMFetch lineitem(orderkey, price, discount) per row; sum revenue per orderkey."),
        ("Q4 PhaseA",  "Orders scan: filter o_orderdate ∈ [1993-07-01, 1993-10-01) → orderkey_priority map."),
        ("Q4 PhaseB",  "OR matching orderkey bitmaps from bitmap_orderkey; get_rowids."),
        ("Q4 PhaseC",  "BMFetch lineitem(orderkey, commitdate, receiptdate); per-row eval EXISTS commit < receipt; collect orderkey set."),
        ("Q4 PhaseD",  "Count by priority for orderkeys in the EXISTS set."),
        ("Q5 PhaseA",  "5 explicit semi-joins: region(ASIA) → nation → customer → orders(date range) → supplier."),
        ("Q5 PhaseB",  "Multi-OR over l_orderkey bitmaps for qualifying orders + multi-OR over l_suppkey bitmaps for qualifying suppliers; AND."),
        ("Q5 PhaseC",  "BMFetch lineitem(orderkey, suppkey, price, discount); aggregate revenue per nation."),
        ("Q6 ship_GE", "Apply per-year shipdate bitmap (Btvs_GE[1994])."),
        ("Q6 OR_disc", "Multi-OR over discount bitmaps in [5, 7], or single AND_NOT via discount_BPE."),
        ("Q6 OR_qty",  "Multi-OR over quantity bitmaps in [0, 2399 raw], or single AND_NOT via quantity_BPE."),
        ("Q6 AND",     "btv_disc &= btv_qty; btv_disc &= btv_ship_GE."),
        ("Q6 GetRowIds", "Extract sorted row IDs from final result bitmap."),
        ("Q14 PhaseA", "Shipdate range OR (1 month: l_shipdate ∈ ['1995-09-01', '1995-10-01'))."),
        ("Q14 PhaseB", "Part scan: p_type LIKE 'PROMO%' → promo_partkeys."),
        ("Q14 PhaseC", "BMFetch lineitem(partkey, price, discount); accumulate total_rev + promo_rev (if l_partkey ∈ promo_set)."),
        ("Q15 OR_ship", "Single-column shipdate range OR ([1996-01-01, 1996-04-01))."),
        ("Q15 walk",   "Walk filter rows; revenue_by_suppkey[col_suppkey[r]] += price * (100 - disc)."),
        ("Q17 PhaseA", "Part scan: p_brand='Brand#23' AND p_container='MED BOX'; per-match OR bitmap_partkey[partkey]."),
        ("Q17 PhaseB", "BMFetch lineitem(partkey, qty, extprice); compute per-partkey avg(qty)."),
        ("Q17 PhaseC", "Filter rows where qty < 0.2 × avg_qty[partkey]; sum extprice."),
        ("Q19 PhaseA", "bitmap_shipmode[A] AND bitmap_shipinstruct[D] (covers AIR + AIR REG + DELIVER IN PERSON)."),
        ("Q19 PhaseB", "Part scan → unordered_map<partkey, GroupSpec> for the 3 OR-groups (brand×container×size×qty range)."),
        ("Q19 PhaseC", "BMFetch lineitem(partkey, qty, price, disc); per-row look up partkey → group; check qty range; sum revenue."),
    ]
    for tag, doc in phase_glossary:
        readme.cell(row=rr, column=1, value=tag).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        rr += 1
    rr += 1

    _section("Algorithm matrix (per backend k-way OR strategy)")
    algo_matrix = [
        ("CB",  "Pairwise SparseComBit::apply_or_to (no native k-way merge; sparse segment storage already amortises common cases)."),
        ("CR",  "Pairwise |= (CRoaring's standard mode — comparison baseline)."),
        ("CRR", "fastunion (priority-queue k-way merge over run-optimised bitmaps; CRoaring's advertised path)."),
        ("WAH", "Pairwise |= (FastBit has no public k-way merge API)."),
        ("EW",  "fast_logicalor (k-way priority-queue merge; EWAH's idiomatic path)."),
        ("CON", "fast_logicalor (k-way merge; Concise's idiomatic path)."),
        ("BS",  "Scalar word-level OR (uncompressed reference, single-pass through 64-bit words)."),
        ("BSA", "AVX-512 word-level OR (same data layout as BS, vectorised)."),
    ]
    for tag, doc in algo_matrix:
        readme.cell(row=rr, column=1, value=tag).font = Q_FONT
        readme.cell(row=rr, column=2, value=doc)
        rr += 1

    readme.column_dimensions["A"].width = 12
    readme.column_dimensions["B"].width = 110

    # ------------------------------------------------------------
    # Sheet: Time Matrix (Q rows × backend cols, conditional colour scale)
    # ------------------------------------------------------------
    ws = wb.create_sheet("Time Matrix")
    _put_header(ws, 1, ["Q"] + BACKENDS + ["DuckDB SQL"])
    for i, q in enumerate(QUERIES, start=2):
        q_tag = f"Q{q}"
        c = ws.cell(row=i, column=1, value=q_tag)
        c.fill = Q_FILL; c.font = Q_FONT; c.alignment = CENTER; c.border = BORDER
        for j, b in enumerate(BACKENDS, start=2):
            v = _q_total_ms(q_tag, b)
            cell = ws.cell(row=i, column=j, value=v)
            cell.number_format = NUM_FMT_MS
            cell.alignment = RIGHT
            cell.border = BORDER
            if v is None:
                cell.fill = NUM_FILL_NA
        v = meta.get(q_tag, {}).get("duckdb_ms")
        cell = ws.cell(row=i, column=2 + len(BACKENDS), value=v)
        cell.number_format = NUM_FMT_MS
        cell.alignment = RIGHT
        cell.border = BORDER
        if v is None:
            cell.fill = NUM_FILL_NA
    # Per-row green→red colour scale across the 8 backend columns.
    last_row = 1 + len(QUERIES)
    for i in range(2, last_row + 1):
        rng = f"{get_column_letter(2)}{i}:{get_column_letter(1+len(BACKENDS))}{i}"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B"))
    _autosize(ws, min_w=12, max_w=18)
    ws.column_dimensions["A"].width = 6

    # ------------------------------------------------------------
    # Sheet: Phase Time Detail
    # ------------------------------------------------------------
    ws = wb.create_sheet("Phase Time Detail")
    _put_header(ws, 1, ["Q", "Phase"] + BACKENDS)
    rownum = 2
    # group records by Q then by Phase (preserve first-seen phase order
    # within a Q so the natural OR_ship → NOT → AND+Agg order shows up).
    seen: Dict[str, List[str]] = {}
    for r in records:
        seen.setdefault(r["Q"], [])
        if r["Phase"] not in seen[r["Q"]]:
            seen[r["Q"]].append(r["Phase"])
    for q in QUERIES:
        q_tag = f"Q{q}"
        phases = seen.get(q_tag, [])
        if not phases:
            continue
        first_row = rownum
        for phase in phases:
            c = ws.cell(row=rownum, column=1, value=q_tag if rownum == first_row else "")
            c.fill = Q_FILL; c.font = Q_FONT; c.alignment = CENTER; c.border = BORDER
            c2 = ws.cell(row=rownum, column=2, value=phase)
            c2.fill = SUB_FILL; c2.font = SUB_FONT; c2.alignment = LEFT; c2.border = BORDER
            for j, b in enumerate(BACKENDS, start=3):
                v = None
                for r in records:
                    if r["Q"] == q_tag and r["Backend"] == b and r["Phase"] == phase:
                        v = r["Median_ms"]
                        # 0.0 sentinel from per-Q CSV → render as blank
                        # (un-run backend, not "instant query").
                        if v == 0.0:
                            v = None
                        break
                cell = ws.cell(row=rownum, column=j, value=v)
                cell.number_format = NUM_FMT_MS
                cell.alignment = RIGHT
                cell.border = BORDER
                if v is None:
                    cell.fill = NUM_FILL_NA
            rng = f"{get_column_letter(3)}{rownum}:{get_column_letter(2+len(BACKENDS))}{rownum}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B"))
            rownum += 1
        # Merge the Q label cells across the phases for visual grouping.
        if rownum - first_row > 1:
            ws.merge_cells(start_row=first_row, start_column=1,
                           end_row=rownum-1, end_column=1)
    _autosize(ws, min_w=11, max_w=18)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16

    # ------------------------------------------------------------
    # Sheet: Memory Breakdown — multi-line cells per (Q, backend)
    # ------------------------------------------------------------
    ws = wb.create_sheet("Memory Breakdown")
    _put_header(ws, 1, ["Q"] + [f"{b} — {BACKEND_FULL[b]}" for b in BACKENDS])
    for i, q in enumerate(QUERIES, start=2):
        q_tag = f"Q{q}"
        bd_map = meta.get(q_tag, {}).get("memory_breakdown", {})
        c = ws.cell(row=i, column=1, value=q_tag)
        c.fill = Q_FILL; c.font = Q_FONT; c.alignment = CENTER; c.border = BORDER
        # Compute row height based on the longest backend's category list.
        max_lines = 0
        for j, b in enumerate(BACKENDS, start=2):
            bd = bd_map.get(b, {})
            cell = ws.cell(row=i, column=j)
            cell.alignment = LEFT
            cell.border = BORDER
            if not bd:
                cell.value = "—"
                cell.fill = NUM_FILL_NA
                continue
            # Render in canonical category order first (legacy
            # `[Breakdown]` shape: L1/L2/L3/L4 etc), then any leftover
            # keys (per-column entries from the new BMTPCH `[load_bitmap]`
            # path: `orderkey`, `shipdate`, `suppkey`, …), then 'total'.
            tag_order = BREAKDOWN_TAG_ORDER_BY_BACKEND.get(b, [])
            seen = set()
            lines = []
            for cat in tag_order:
                if cat in bd:
                    lines.append(f"{cat}: {bd[cat]:.2f} MB")
                    seen.add(cat)
            for cat, mb in bd.items():
                if cat in seen or cat == "total":
                    continue
                lines.append(f"{cat}: {mb:.2f} MB")
            if "total" in bd:
                lines.append(f"total: {bd['total']:.2f} MB")
            cell.value = "\n".join(lines)
            max_lines = max(max_lines, len(lines))
        # 14pt per line + a little headroom
        ws.row_dimensions[i].height = max(36, 14 * max(1, max_lines) + 6)
    _autosize(ws, min_w=18, max_w=34)
    ws.column_dimensions["A"].width = 6

    # ------------------------------------------------------------
    # Sheet: Memory Detail (MB) — long format for plots
    # ------------------------------------------------------------
    ws = wb.create_sheet("Memory Detail (MB)")
    _put_header(ws, 1, ["Q", "Backend", "Backend (full)", "Category", "MB"])
    rownum = 2
    for q in QUERIES:
        q_tag = f"Q{q}"
        bd_map = meta.get(q_tag, {}).get("memory_breakdown", {})
        for b in BACKENDS:
            bd = bd_map.get(b, {})
            if not bd:
                continue
            tag_order = BREAKDOWN_TAG_ORDER_BY_BACKEND.get(b, [])
            for cat in tag_order + ["total"]:
                if cat not in bd:
                    continue
                ws.cell(row=rownum, column=1, value=q_tag)
                ws.cell(row=rownum, column=2, value=b)
                ws.cell(row=rownum, column=3, value=BACKEND_FULL[b])
                ws.cell(row=rownum, column=4, value=cat)
                cell = ws.cell(row=rownum, column=5, value=bd[cat])
                cell.number_format = NUM_FMT_MB
                rownum += 1
    _autosize(ws, min_w=10, max_w=22)

    # ------------------------------------------------------------
    # Sheet: Storage Total (MB)
    # ------------------------------------------------------------
    ws = wb.create_sheet("Storage Total (MB)")
    _put_header(ws, 1, ["Q"] + BACKENDS + ["On-disk Σ", "In-mem Σ", "Note"])
    for i, q in enumerate(QUERIES, start=2):
        q_tag = f"Q{q}"
        m = meta.get(q_tag, {})
        c = ws.cell(row=i, column=1, value=q_tag)
        c.fill = Q_FILL; c.font = Q_FONT; c.alignment = CENTER; c.border = BORDER
        disk_sum = 0.0
        mem_sum  = 0.0
        any_shared = False
        for j, b in enumerate(BACKENDS, start=2):
            mb = m.get("footprint_mb", {}).get(b)
            cell = ws.cell(row=i, column=j, value=mb)
            cell.number_format = NUM_FMT_MB
            cell.alignment = RIGHT
            cell.border = BORDER
            if mb is None:
                cell.fill = NUM_FILL_NA
                continue
            kind = m.get("footprint_kind", {}).get(b) or ""
            if kind.startswith("disk"):
                disk_sum += mb
                if "shared" in kind:
                    any_shared = True
            elif kind == "mem":
                mem_sum += mb
        # Dedupe shared CR/CRR disk and BS/BSA mem.
        cr_mb = m.get("footprint_mb", {}).get("CR")
        crr_kind = m.get("footprint_kind", {}).get("CRR") or ""
        if cr_mb is not None and crr_kind.startswith("disk"):
            disk_sum -= cr_mb
        bs_mb = m.get("footprint_mb", {}).get("BS")
        if bs_mb is not None and m.get("footprint_kind", {}).get("BSA") == "mem":
            mem_sum -= bs_mb
        ws.cell(row=i, column=2 + len(BACKENDS),
                value=round(disk_sum, 2) if disk_sum else None
               ).number_format = NUM_FMT_MB
        ws.cell(row=i, column=3 + len(BACKENDS),
                value=round(mem_sum, 2) if mem_sum else None
               ).number_format = NUM_FMT_MB
        ws.cell(row=i, column=4 + len(BACKENDS),
                value=("symlinked share with another Q" if any_shared else ""))
    _autosize(ws, min_w=11, max_w=22)
    ws.column_dimensions["A"].width = 6

    # ------------------------------------------------------------
    # Sheet: Correctness
    # ------------------------------------------------------------
    ws = wb.create_sheet("Correctness")
    _put_header(ws, 1, ["Q"] + [f"{b} rows" for b in BACKENDS]
                + ["DuckDB SQL", "Validation message"])
    for i, q in enumerate(QUERIES, start=2):
        q_tag = f"Q{q}"
        m = meta.get(q_tag, {})
        c = ws.cell(row=i, column=1, value=q_tag)
        c.fill = Q_FILL; c.font = Q_FONT; c.alignment = CENTER; c.border = BORDER
        for j, b in enumerate(BACKENDS, start=2):
            v = m.get("rows", {}).get(b)
            cell = ws.cell(row=i, column=j, value=v)
            cell.number_format = NUM_FMT_INT
            cell.alignment = RIGHT
            cell.border = BORDER
            if v is None:
                cell.fill = NUM_FILL_NA
        ws.cell(row=i, column=2 + len(BACKENDS),
                value=m.get("duckdb_rows")).number_format = NUM_FMT_INT
        msg_cell = ws.cell(row=i, column=3 + len(BACKENDS),
                           value=m.get("ok_message") or "")
        msg_cell.alignment = LEFT
    _autosize(ws, min_w=12, max_w=80)
    ws.column_dimensions["A"].width = 6

    # Freeze every sheet's header row for convenient scrolling.
    for s in wb.worksheets:
        if s.title != "README":
            s.freeze_panes = "B2"

    wb.save(out_xlsx)


# --- Main -----------------------------------------------------------------


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--sf", type=int, default=10, choices=[1, 10],
                    help="TPC-H scale factor (default: 10)")
    ap.add_argument("--duckdb", type=Path, default=DEFAULT_DUCKDB_BIN,
                    help="Path to the duckdb binary (default: build/release/duckdb)")
    ap.add_argument("--db", type=Path, default=None,
                    help="Path to the TPC-H duckdb DB file (default: tpch_sf{N}.db)")
    ap.add_argument("--out-csv", type=Path,
                    default=REPO_ROOT / "bm_results_long.csv",
                    help="Output long-format CSV path")
    ap.add_argument("--out-xlsx", type=Path,
                    default=REPO_ROOT / "bm_results.xlsx",
                    help="Output Excel workbook path")
    ap.add_argument("--log-dir", type=Path,
                    default=REPO_ROOT / "bm_logs",
                    help="Directory to store per-Q stdout logs")
    ap.add_argument("--skip-run", action="store_true",
                    help="Skip running queries; reparse existing CSVs + logs")
    ap.add_argument("--only", type=int, nargs="+",
                    help="Only run a subset of Q numbers (for debugging)")
    args = ap.parse_args()

    sf_label = f"SF{args.sf}"
    db_path = args.db or REPO_ROOT / DEFAULT_DB_TEMPLATE.format(sf=args.sf)
    args.log_dir.mkdir(parents=True, exist_ok=True)

    # --- Run ---------------------------------------------------------------
    queries = args.only or QUERIES
    if not args.skip_run:
        print(f"[bench] duckdb = {args.duckdb}")
        print(f"[bench] db     = {db_path}")
        print(f"[bench] SF     = {args.sf}")
        print(f"[bench] Qs     = {queries}")
        if not args.duckdb.exists():
            sys.exit(f"[error] duckdb binary not found: {args.duckdb}")
        if not db_path.exists():
            sys.exit(f"[error] DB file not found: {db_path}")
        for q in queries:
            log_path = args.log_dir / f"q{q}_{sf_label}.log"
            print(f"[bench] Q{q} -> {log_path} ...")
            rc = run_single_query(
                args.duckdb, db_path, q,
                env_extra={"TPCH_SF": str(args.sf), "TPCH_SF_LABEL": sf_label},
                log_path=log_path,
            )
            if rc != 0:
                print(f"[warn] Q{q} exited rc={rc}; continuing", file=sys.stderr)

    # --- Parse + aggregate -------------------------------------------------
    records: List[Dict] = []
    meta: Dict[str, Dict] = {}
    for q in queries:
        q_tag = f"Q{q}"
        records.extend(parse_csv_for_q(REPO_ROOT, q, sf_label))
        log_path = args.log_dir / f"q{q}_{sf_label}.log"
        if log_path.exists():
            meta[q_tag] = parse_stdout_log(
                log_path.read_text(errors="replace"),
                q=q, repo_root=REPO_ROOT, sf=args.sf,
            )
        else:
            meta[q_tag] = {"footprint_mb": {}, "footprint_kind": {},
                           "rows": {}, "duckdb_rows": None, "duckdb_ms": None,
                           "ok_message": None}

    print(f"[bench] {len(records)} timing rows parsed across {len(queries)} Qs")

    # --- Emit --------------------------------------------------------------
    write_long_csv(records, args.out_csv)
    print(f"[bench] wrote {args.out_csv}")
    build_excel(records, meta, sf_label, args.out_xlsx)
    print(f"[bench] wrote {args.out_xlsx}")


if __name__ == "__main__":
    main()
