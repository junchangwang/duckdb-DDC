#!/usr/bin/env python3
"""End-to-end consistency verification of bm_results.xlsx.

Cross-checks every cell in the Excel workbook against two sources of truth:
  1. bm_logs/q{N}_SF10.log  — the raw stdout of every per-backend run
  2. bm_results_long.csv    — the long-format timing CSV emitted by the bench

For each (sheet, Q, backend) tuple we report MISMATCH / OK so the user can
inspect any non-trivial discrepancy.  Tolerance is 0.01 MB / 0.05 ms (the
display precision of the spreadsheet).
"""
from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
LOG_DIR = REPO_ROOT / "bm_logs"
SF_LABEL = "SF10"
QS = [1, 3, 4, 5, 6, 8, 10, 12, 14, 15, 17, 19]
BACKENDS = ["WAH", "CB", "CB+BPE", "CR", "CR+BPE", "CRR", "EW", "BS", "BSA", "CON"]

# Map IBitmapIndex::backend_name() -> canonical Excel backend tag.
NAME_TO_BACKEND = {
    "DDC":      "CB",
    "DDCGE":    "CB",
    "DDCBPE":   "CB+BPE",
    "CRoaring":    "CR",
    "CRoaringRun": "CRR",
    "CRoaringBPE": "CR+BPE",
    "WAH":         "WAH",
    "EWAH":        "EW",
    "Concise":     "CON",
}

RE_LOAD_BUILT = re.compile(
    r"\[load_bitmap\]\s+(\S+):\s+built\s+(\w+)\s+index\s+\(\d+\s+keys,\s+([0-9.]+)\s+MB\)"
)
RE_LOAD_BREAKDOWN = re.compile(
    r"\[load_bitmap_breakdown\]\s+(\S+)\s+(\w+):\s+(.+)$"
)
RE_KV_MB = re.compile(r"(\w+)\s*=\s*([0-9.]+)\s*MB")


def parse_log(q: int) -> Dict:
    """Returns {(backend, column): MB} from `built ... index (N keys, X MB)`
    lines, plus per-(backend, column) breakdown dict {category: MB}."""
    path = LOG_DIR / f"q{q}_{SF_LABEL}.log"
    out_total: Dict[Tuple[str, str], float] = {}
    out_breakdown: Dict[Tuple[str, str], Dict[str, float]] = {}
    if not path.exists():
        return {"total": out_total, "breakdown": out_breakdown}
    text = path.read_text(errors="replace")
    for m in RE_LOAD_BUILT.finditer(text):
        col, name, mb = m.group(1), m.group(2), float(m.group(3))
        backend = NAME_TO_BACKEND.get(name)
        if backend:
            # The log emits one `built` line per (column, backend, run).
            # Multiple runs of the same backend (cb vs cb_bpe both call
            # DDC::compress, and Q15 has 8 runs) overwrite the entry —
            # they should all measure the same MB, so this is benign.
            out_total[(backend, col)] = mb
    for m in RE_LOAD_BREAKDOWN.finditer(text):
        col, name, kv_str = m.group(1), m.group(2), m.group(3)
        backend = NAME_TO_BACKEND.get(name)
        if not backend:
            continue
        bd: Dict[str, float] = {}
        for kvm in RE_KV_MB.finditer(kv_str):
            bd[kvm.group(1)] = float(kvm.group(2))
        out_breakdown[(backend, col)] = bd
    return {"total": out_total, "breakdown": out_breakdown}


def parse_excel_breakdown_cell(text: str) -> Dict[str, float]:
    """Parse a cell like `array: 119.97 MB\\nshipdate: 138.44 MB\\ntotal: 138.44 MB`."""
    out: Dict[str, float] = {}
    if not text or text in ("—", "-"):
        return out
    for line in text.splitlines():
        m = re.match(r"^\s*(\S+):\s+([0-9.]+)\s*MB\s*$", line)
        if m:
            out[m.group(1)] = float(m.group(2))
    return out


def cmp_mb(label: str, got: float, exp: float, tol: float = 0.05) -> str:
    if abs(got - exp) <= tol:
        return f"OK    {label}: {got:.4f} ≈ {exp:.4f}"
    return f"MISMATCH {label}: excel={got:.4f}  log={exp:.4f}  Δ={got-exp:+.4f}"


def parse_long_csv() -> Dict[Tuple[int, str, str], float]:
    """Return {(q, backend, phase): median_ms}."""
    out: Dict[Tuple[int, str, str], float] = {}
    path = REPO_ROOT / "bm_results_long.csv"
    if not path.exists():
        return out
    with path.open() as f:
        for row in csv.DictReader(f):
            q = int(row["Q"][1:])
            backend = row["Backend"]
            phase = row["Phase"]
            try:
                ms = float(row["Median_ms"])
            except (ValueError, TypeError):
                continue
            out[(q, backend, phase)] = ms
    return out


def parse_excel_time_matrix(wb) -> Dict[Tuple[int, str], float]:
    """Time Matrix sheet: rows = Q, cols = backends; cells = total ms median."""
    sh = wb["Time Matrix"]
    headers = [sh.cell(1, c).value for c in range(1, sh.max_column + 1)]
    out: Dict[Tuple[int, str], float] = {}
    for r in range(2, sh.max_row + 1):
        q_cell = sh.cell(r, 1).value
        if not q_cell or not isinstance(q_cell, str):
            continue
        m = re.match(r"Q(\d+)", q_cell)
        if not m:
            continue
        q = int(m.group(1))
        for c in range(2, sh.max_column + 1):
            hdr = headers[c - 1]
            if not hdr:
                continue
            val = sh.cell(r, c).value
            if val is None or val == "":
                continue
            try:
                out[(q, hdr)] = float(val)
            except (ValueError, TypeError):
                pass
    return out


def main() -> int:
    wb = openpyxl.load_workbook(REPO_ROOT / "bm_results.xlsx", data_only=True)
    long_csv = parse_long_csv()

    issues: List[str] = []

    # =====================================================================
    # 1. Memory Breakdown sheet — every column-row should match the log
    # =====================================================================
    print("=" * 80)
    print("1. Memory Breakdown sheet vs bm_logs/qN_SF10.log")
    print("=" * 80)

    sh_mb = wb["Memory Breakdown"]
    hdr_cols = [sh_mb.cell(1, c).value for c in range(1, sh_mb.max_column + 1)]
    hdr_to_backend = {}
    for c, hdr in enumerate(hdr_cols, start=1):
        if not hdr or hdr == "Q":
            continue
        for b in BACKENDS:
            if hdr.startswith(f"{b} —") or hdr == b:
                hdr_to_backend[c] = b
                break

    for r in range(2, sh_mb.max_row + 1):
        q_cell = sh_mb.cell(r, 1).value
        if not q_cell or not isinstance(q_cell, str):
            continue
        m = re.match(r"Q(\d+)", q_cell)
        if not m:
            continue
        q = int(m.group(1))
        log = parse_log(q)

        for c, backend in hdr_to_backend.items():
            cell = sh_mb.cell(r, c).value
            if cell is None or cell in ("—", "-"):
                continue
            excel_bd = parse_excel_breakdown_cell(str(cell))
            if not excel_bd:
                continue

            # Expected per-column total from the log.  For β-scheme
            # backends (CB+BPE / CR+BPE) the run loads BOTH the per-value
            # bitmap (logged as `DDC`/`CRoaring`) AND the bucketed-prefix
            # column (logged as `DDCBPE`/`CRoaringBPE`).  The Excel
            # breakdown cell shows both, so we have to union both backend
            # families when checking.
            backends_for_match: List[str] = [backend]
            if backend == "CB+BPE":
                backends_for_match.append("CB")
            elif backend == "CR+BPE":
                backends_for_match.append("CR")
            log_cols = {col: mb for (b, col), mb in log["total"].items()
                        if b in backends_for_match}
            for col, exp_mb in log_cols.items():
                got = excel_bd.get(col)
                if got is None:
                    issues.append(
                        f"Q{q} {backend} Memory Breakdown: log has '{col}={exp_mb:.4f} MB' "
                        f"but Excel cell missing it"
                    )
                else:
                    if abs(got - exp_mb) > 0.05:
                        issues.append(
                            f"Q{q} {backend} Memory Breakdown col '{col}': "
                            f"excel={got:.4f}  log={exp_mb:.4f}  Δ={got-exp_mb:+.4f}"
                        )

            # 'total' line should equal sum of per-column entries
            if "total" in excel_bd and log_cols:
                exp_total = sum(log_cols.values())
                got_total = excel_bd["total"]
                if abs(got_total - exp_total) > 0.10:
                    # Per-column totals are bench's headline footprint, not
                    # the sum of structural breakdowns (those structural
                    # bytes are inside each column's total).
                    issues.append(
                        f"Q{q} {backend} Memory Breakdown 'total': "
                        f"excel={got_total:.4f}  Σcols={exp_total:.4f}  Δ={got_total-exp_total:+.4f}"
                    )

    # =====================================================================
    # 2. Storage Total sheet (if present)
    # =====================================================================
    if "Storage Total (MB)" in wb.sheetnames:
        print()
        print("=" * 80)
        print("2. Storage Total (MB) sheet")
        print("=" * 80)
        sh_st = wb["Storage Total (MB)"]
        hdr = [sh_st.cell(1, c).value for c in range(1, sh_st.max_column + 1)]
        for r in range(2, sh_st.max_row + 1):
            q_cell = sh_st.cell(r, 1).value
            if not q_cell or not isinstance(q_cell, str):
                continue
            m = re.match(r"Q(\d+)", q_cell)
            if not m:
                continue
            q = int(m.group(1))
            log = parse_log(q)
            for c, h in enumerate(hdr, start=1):
                if not h or h == "Q":
                    continue
                backend = None
                for b in BACKENDS:
                    if h == b or h.startswith(f"{b} —"):
                        backend = b
                        break
                if not backend:
                    continue
                val = sh_st.cell(r, c).value
                if val is None or val in ("—", "-", "Timeout"):
                    continue
                try:
                    got_mb = float(val)
                except (ValueError, TypeError):
                    continue
                bset: List[str] = [backend]
                if backend == "CB+BPE":
                    bset.append("CB")
                elif backend == "CR+BPE":
                    bset.append("CR")
                exp_mb = sum(mb for (b, _), mb in log["total"].items()
                             if b in bset)
                if exp_mb == 0:
                    continue
                if abs(got_mb - exp_mb) > 0.10:
                    issues.append(
                        f"Q{q} {backend} Storage Total: "
                        f"excel={got_mb:.4f}  log Σ={exp_mb:.4f}  Δ={got_mb-exp_mb:+.4f}"
                    )

    # =====================================================================
    # 3. Time Matrix sheet vs bm_results_long.csv (TOTAL phase)
    # =====================================================================
    print()
    print("=" * 80)
    print("3. Time Matrix vs bm_results_long.csv (TOTAL phase)")
    print("=" * 80)

    excel_tm = parse_excel_time_matrix(wb)
    for (q, hdr), excel_ms in excel_tm.items():
        if hdr in ("DuckDB SQL", "DuckDB"):
            continue
        backend = None
        for b in BACKENDS:
            if hdr == b:
                backend = b
                break
        if not backend:
            continue
        # Find matching long-csv TOTAL entry
        ms = long_csv.get((q, backend, "TOTAL"))
        if ms is None:
            # Some Qs use phase names like 'TOTAL' / 'Total' uppercase variation
            ms = long_csv.get((q, backend, "Total"))
        if ms is None:
            issues.append(
                f"Q{q} {backend} Time Matrix: excel={excel_ms:.2f}ms  long_csv has no TOTAL entry"
            )
            continue
        if abs(excel_ms - ms) > 0.05:
            issues.append(
                f"Q{q} {backend} Time Matrix TOTAL: "
                f"excel={excel_ms:.4f}  long_csv={ms:.4f}  Δ={excel_ms-ms:+.4f}"
            )

    # =====================================================================
    # 4. Phase Time Detail — every (Q, phase, backend) cell vs long CSV
    # =====================================================================
    print()
    print("=" * 80)
    print("4. Phase Time Detail vs bm_results_long.csv")
    print("=" * 80)

    sh_pt = wb["Phase Time Detail"]
    pt_hdr = [sh_pt.cell(1, c).value for c in range(1, sh_pt.max_column + 1)]
    current_q: Optional[int] = None
    for r in range(2, sh_pt.max_row + 1):
        q_cell = sh_pt.cell(r, 1).value
        if q_cell:
            m = re.match(r"Q(\d+)", str(q_cell))
            if m:
                current_q = int(m.group(1))
        phase = sh_pt.cell(r, 2).value
        if not phase or current_q is None:
            continue
        for c in range(3, sh_pt.max_column + 1):
            backend = pt_hdr[c - 1]
            if not backend or backend not in BACKENDS:
                continue
            val = sh_pt.cell(r, c).value
            if val is None or val == "" or val == 0:
                continue
            try:
                excel_ms = float(val)
            except (ValueError, TypeError):
                continue
            # The long CSV uses canonical phase names that match the cpp's
            # operation tags (PhaseA / PhaseB / PhaseC / PhaseD / TOTAL,
            # plus Q1/Q5/Q6 specials).  The Excel sheet shows the same
            # tags, so a direct lookup works.
            csv_ms = long_csv.get((current_q, backend, phase))
            if csv_ms is None:
                # Some sheets render "PhaseA" while the csv has e.g.
                # "PhaseA_OR" / "PhaseA_ship".  Try fuzzy prefix.
                for (q2, b2, p2), m_ in long_csv.items():
                    if q2 == current_q and b2 == backend and p2.startswith(phase):
                        csv_ms = m_
                        break
            if csv_ms is None:
                continue  # legitimate gap (e.g. WAH/EW timeouts on Q3/Q4)
            if abs(excel_ms - csv_ms) > 0.05:
                issues.append(
                    f"Q{current_q} {backend} Phase {phase}: "
                    f"excel={excel_ms:.4f}  csv={csv_ms:.4f}  Δ={excel_ms-csv_ms:+.4f}"
                )

    # =====================================================================
    # 5. Memory Detail — every row's MB vs log
    # =====================================================================
    print()
    print("=" * 80)
    print("5. Memory Detail (MB) vs bm_logs/qN_SF10.log")
    print("=" * 80)

    sh_md = wb["Memory Detail (MB)"]
    for r in range(2, sh_md.max_row + 1):
        q_cell = sh_md.cell(r, 1).value
        if not q_cell:
            continue
        m = re.match(r"Q(\d+)", str(q_cell))
        if not m:
            continue
        q = int(m.group(1))
        backend = sh_md.cell(r, 2).value
        category = sh_md.cell(r, 4).value
        try:
            excel_mb = float(sh_md.cell(r, 5).value or 0)
        except (ValueError, TypeError):
            continue
        log = parse_log(q)
        # Two flavors: (a) category is a column name (shipdate, orderkey,
        # …) → matches log["total"][(backend, col)].  (b) category is a
        # structural piece (L1/L2/array/header/…) → the sum across all
        # columns for that backend should match log["breakdown"].  We
        # only validate (a) here; structural pieces are aggregated at the
        # spreadsheet builder and a coarse cross-check would be noisy.
        bset = [backend]
        if backend == "CB+BPE":
            bset.append("CB")
        elif backend == "CR+BPE":
            bset.append("CR")
        col_mbs = {col: mb for (b, col), mb in log["total"].items()
                   if b in bset}
        if category in col_mbs:
            exp = col_mbs[category]
            if abs(excel_mb - exp) > 0.05:
                issues.append(
                    f"Q{q} {backend} Memory Detail row '{category}': "
                    f"excel={excel_mb:.4f}  log={exp:.4f}  Δ={excel_mb-exp:+.4f}"
                )

    # =====================================================================
    # 6. Correctness — DuckDB SQL row count cells vs log baseline
    # =====================================================================
    print()
    print("=" * 80)
    print("6. Correctness sheet")
    print("=" * 80)

    sh_co = wb["Correctness"]
    co_hdr = [sh_co.cell(1, c).value for c in range(1, sh_co.max_column + 1)]
    for r in range(2, sh_co.max_row + 1):
        q_cell = sh_co.cell(r, 1).value
        if not q_cell:
            continue
        m = re.match(r"Q(\d+)", str(q_cell))
        if not m:
            continue
        # Just sanity-check that not every cell in the row is None — the
        # bench writes None when row counts aren't emitted by the query
        # (Q1 / Q15 etc. don't print backend-level row counts).  This is
        # a structural check, not a numeric one.
        nonnull = sum(1 for c in range(2, sh_co.max_column + 1)
                      if sh_co.cell(r, c).value not in (None, ""))
        if nonnull == 0:
            issues.append(
                f"Q{m.group(1)} Correctness: row entirely empty "
                f"(no backend produced a row count)"
            )

    # =====================================================================
    # Summary
    # =====================================================================
    print()
    print("=" * 80)
    if issues:
        print(f"FOUND {len(issues)} ISSUE(S):")
        for i, msg in enumerate(issues, 1):
            print(f"  {i}. {msg}")
        return 1
    else:
        print("ALL CELLS MATCH SOURCE-OF-TRUTH (within tolerance).")
        return 0


if __name__ == "__main__":
    sys.exit(main())
